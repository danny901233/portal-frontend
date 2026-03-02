"""
agent_infra.py — Tyresoft ReceptionMate Infrastructure
Constants, speech formatting, VRN handling, sanitisation, API client,
tyre inventory, specialist LLMs, error monitoring, dynamic greeting.
"""

import os
import re
import csv
import json
import time
import asyncio
import base64
import datetime
from dataclasses import dataclass, field
from enum import Enum
from typing import Optional, Any
from pathlib import Path
from zoneinfo import ZoneInfo

import aiohttp
from dotenv import load_dotenv

try:
    from openai import AsyncOpenAI
    HAS_OPENAI = True
except ImportError:
    AsyncOpenAI = None
    HAS_OPENAI = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

load_dotenv(".env.local")

# ═══════════════════════════════════════════════════════════════════════════
# TIMEZONE
# ═══════════════════════════════════════════════════════════════════════════
UK_TZ = ZoneInfo("Europe/London")


def uk_now() -> datetime.datetime:
    return datetime.datetime.now(UK_TZ)


def uk_timestamp() -> str:
    return uk_now().strftime("%Y-%m-%d %H:%M:%S")


def uk_date() -> str:
    return uk_now().strftime("%Y-%m-%d")


# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════
ELEVEN_VOICE_ID = os.getenv("ELEVEN_VOICE_ID", "21m00Tcm4TlvDq8ikWAM")
ELEVEN_TTS_MODEL = os.getenv("ELEVEN_TTS_MODEL", "eleven_turbo_v2_5")
ELEVEN_STABILITY = float(os.getenv("ELEVEN_STABILITY", "0.45"))
ELEVEN_SIMILARITY = float(os.getenv("ELEVEN_SIMILARITY", "0.78"))
ELEVEN_STYLE = float(os.getenv("ELEVEN_STYLE", "0.35"))

TYRESOFT_WORKSPACE = os.getenv("TYRESOFT_WORKSPACE", "test")
TYRESOFT_API_USERNAME = os.getenv("TYRESOFT_API_USERNAME")
TYRESOFT_API_PASSWORD = os.getenv("TYRESOFT_API_PASSWORD")
TYRESOFT_X_API_KEY = os.getenv("TYRESOFT_X_API_KEY")
TYRESOFT_BASE_URL = f"https://3p-api.tyresoft.biz/v1/{TYRESOFT_WORKSPACE}"
API_TIMEOUT = aiohttp.ClientTimeout(total=30, connect=10)

SPEAKING_MODEL = os.getenv("SPEAKING_MODEL", "openai/gpt-4o-mini")
SPECIALIST_MODEL = os.getenv("SPECIALIST_MODEL", "openai/gpt-4o-mini")
SPECIALIST_TIMEOUT = 5.0  # seconds — 3s was too short, cold starts take 1-2s

DISCORD_WEBHOOK_URL = os.getenv("DISCORD_WEBHOOK_URL", "")
ERROR_LOG_PATH = os.getenv(
    "ERROR_LOG_PATH", os.path.join(os.path.dirname(__file__), "error_log.xlsx")
)
AGENT_BRANCH_NAME = os.getenv("AGENT_BRANCH_NAME", "Tyresoft")

# LiveKit Inference Gateway
_LIVEKIT_INFERENCE_URL = "https://agent-gateway.livekit.cloud/v1"
_specialist_llm: Optional[Any] = None

CHANNEL_ID = 24  # Reception Mate API channel


# ═══════════════════════════════════════════════════════════════════════════
# SERVICES & BRANCHES
# ═══════════════════════════════════════════════════════════════════════════
SERVICES = {
    "WA":    {"name": "Wheel Alignment",                "price": 47.99,  "service_id": 3,  "engine_from": 0,    "engine_to": 0,    "fuel": ""},
    "AIR1":  {"name": "Air Con Recharge - R134a",       "price": 84.00,  "service_id": 11, "engine_from": 0,    "engine_to": 0,    "fuel": ""},
    "FS1":   {"name": "Full Service 0cc-1199cc",        "price": 132.00, "service_id": 2,  "engine_from": 0,    "engine_to": 1199, "fuel": "Diesel or Petrol"},
    "FS2":   {"name": "Full Service 1200cc-1599cc",     "price": 156.00, "service_id": 57, "engine_from": 1200, "engine_to": 1599, "fuel": "Diesel or Petrol"},
    "FS3":   {"name": "Full Service 1600cc-1999cc",     "price": 175.00, "service_id": 8,  "engine_from": 1600, "engine_to": 99999, "fuel": "Diesel or Petrol"},
    "FSE1":  {"name": "Hybrid Vehicle Service",         "price": 155.94, "service_id": 20, "engine_from": 0,    "engine_to": 0,    "fuel": "Hybrid"},
    "MOT-4": {"name": "MOT Class 4 Online",             "price": 50.00,  "service_id": 58, "engine_from": 0,    "engine_to": 0,    "fuel": ""},
    "PUNC":  {"name": "Puncture Repair",                "price": 12.00,  "service_id": 22, "engine_from": 0,    "engine_to": 0,    "fuel": ""},
}

BRANCHES = {
    1: {"name": "Test Auto Service (Branch 1)", "depot_id": 1},
    2: {"name": "Test Auto Service (Branch 2)", "depot_id": 3},
}


def match_full_service(engine_cc: int, fuel: str = "") -> Optional[str]:
    """Match engine size + fuel to correct Full Service code."""
    fuel_lower = fuel.lower() if fuel else ""
    if "hybrid" in fuel_lower or "electric" in fuel_lower:
        return "FSE1"
    for code in ("FS1", "FS2", "FS3"):
        svc = SERVICES[code]
        if svc["engine_from"] <= engine_cc <= svc["engine_to"]:
            return code
    return "FS3"  # default to largest tier


# ═══════════════════════════════════════════════════════════════════════════
# CONSTANTS — NATO, DIGITS, B/V/P
# ═══════════════════════════════════════════════════════════════════════════
_NATO_LETTER_MAP = {
    "alpha": "A", "bravo": "B", "charlie": "C", "delta": "D", "david": "D",
    "echo": "E", "edward": "E", "foxtrot": "F", "freddy": "F",
    "golf": "G", "george": "G", "hotel": "H", "harry": "H", "henry": "H",
    "india": "I", "juliet": "J", "kilo": "K", "king": "K",
    "lima": "L", "london": "L", "mike": "M", "mary": "M",
    "november": "N", "nancy": "N", "oscar": "O", "oliver": "O",
    "papa": "P", "peter": "P", "quebec": "Q", "queen": "Q",
    "romeo": "R", "robert": "R", "roger": "R", "sierra": "S", "samuel": "S",
    "tango": "T", "tommy": "T", "uniform": "U", "uncle": "U",
    "victor": "V", "victory": "V", "victoria": "V",
    "whiskey": "W", "william": "W", "xray": "X", "yankee": "Y", "yellow": "Y",
    "zulu": "Z", "zebra": "Z",
}

_DIGIT_WORD_MAP = {
    "zero": "0", "oh": "0", "owe": "0", "o": "0", "naught": "0",
    "one": "1", "won": "1", "two": "2", "too": "2",
    "three": "3", "four": "4", "for": "4",
    "five": "5", "six": "6", "seven": "7",
    "eight": "8", "ate": "8", "nine": "9", "niner": "9",
}

_BVP_SWAPS = {"B": ["V", "P"], "V": ["B", "P"], "P": ["B", "V"]}

_CAR_MAKE_MODEL_WORDS = {
    "land", "rover", "range", "landrover", "bmw", "audi", "ford", "honda",
    "toyota", "nissan", "mazda", "suzuki", "vauxhall", "volkswagen", "vw",
    "mercedes", "benz", "peugeot", "renault", "citroen", "fiat", "alfa",
    "romeo", "seat", "skoda", "hyundai", "kia", "volvo", "saab", "mini",
    "jaguar", "lexus", "subaru", "mitsubishi", "chrysler", "dodge", "jeep",
    "chevrolet", "tesla", "porsche", "ferrari", "lamborghini", "maserati",
    "bentley", "rolls", "royce", "aston", "martin", "hatchback", "saloon",
    "estate", "suv", "coupe", "convertible", "cabriolet",
}

_PLUS_MISHEARINGS = ("flush", "blush", "plush", "plus", "flash")

_SPOKEN_DIGITS = {
    "zero": "0", "one": "1", "two": "2", "three": "3", "four": "4",
    "five": "5", "six": "6", "seven": "7", "eight": "8", "nine": "9",
}


# ═══════════════════════════════════════════════════════════════════════════
# STEP ENUM + CALLSTATE
# ═══════════════════════════════════════════════════════════════════════════
class Step(Enum):
    GREETING = "greeting"
    NEED_VRN = "need_vrn"
    CONFIRMING_VEHICLE = "confirming_vehicle"
    BUILDING_BASKET = "building_basket"
    NEED_TIMESLOT = "need_timeslot"
    NEED_CONTACT = "need_contact"
    CONFIRMED = "confirmed"
    DONE = "done"
    MESSAGE_ONLY = "message_only"


@dataclass
class CallState:
    step: Step = Step.GREETING
    intent: str = ""  # "tyre_purchase", "service_booking", "combined", "message"

    # Caller
    customer_name_first: str = ""
    customer_name_last: str = ""
    contact_phone: str = ""
    contact_email: str = ""

    # VRN
    vrn: str = ""
    vrn_confirmed: bool = False
    vrn_pending: str = ""
    vrn_partial: str = ""
    vrn_attempts: int = 0

    # Vehicle
    vehicle_info: dict = field(default_factory=dict)
    vehicle_make: str = ""
    vehicle_model: str = ""
    vehicle_year: str = ""
    vehicle_engine_cc: str = ""
    vehicle_fuel: str = ""
    tyre_size_options: list = field(default_factory=list)
    selected_tyre_size: str = ""
    selected_tyre_search_string: str = ""

    # Search
    last_search_results: list = field(default_factory=list)

    # Basket
    basket_items: list = field(default_factory=list)

    # Booking
    selected_branch: int = 1
    available_slots: list = field(default_factory=list)
    booking_date: str = ""
    booking_time: str = ""
    selected_slot: dict = field(default_factory=dict)
    booking_submit_pending: bool = False
    timeslot_attempts: int = 0

    # API IDs
    customer_id: int = 0
    vehicle_id: int = 0

    # Transcripts
    recent_transcripts: list = field(default_factory=list)

    # Session
    call_ended: bool = False
    room_name: str = ""


# ═══════════════════════════════════════════════════════════════════════════
# SPEECH FORMATTING
# ═══════════════════════════════════════════════════════════════════════════
def format_vrm_for_speech(vrm: str) -> str:
    """R V Zero Six L N T — spaced readback for TTS."""
    nato_out = {
        'A': 'Alpha', 'B': 'Bravo', 'C': 'Charlie', 'D': 'Delta', 'E': 'Echo',
        'F': 'Foxtrot', 'G': 'Golf', 'H': 'Hotel', 'I': 'India', 'J': 'Juliet',
        'K': 'Kilo', 'L': 'Lima', 'M': 'Mike', 'N': 'November', 'O': 'Oscar',
        'P': 'Papa', 'Q': 'Quebec', 'R': 'Romeo', 'S': 'Sierra', 'T': 'Tango',
        'U': 'Uniform', 'V': 'Victor', 'W': 'Whiskey', 'X': 'X-ray',
        'Y': 'Yankee', 'Z': 'Zulu',
        '0': 'Zero', '1': 'One', '2': 'Two', '3': 'Three', '4': 'Four',
        '5': 'Five', '6': 'Six', '7': 'Seven', '8': 'Eight', '9': 'Nine',
    }
    return "  ".join(nato_out.get(c, c) for c in vrm.upper())


def format_price_for_speech(price: float) -> str:
    ones = ['', 'one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight',
            'nine', 'ten', 'eleven', 'twelve', 'thirteen', 'fourteen', 'fifteen',
            'sixteen', 'seventeen', 'eighteen', 'nineteen']
    tens = ['', '', 'twenty', 'thirty', 'forty', 'fifty', 'sixty', 'seventy',
            'eighty', 'ninety']

    def _num_words(n: int) -> str:
        if n < 20:
            return ones[n]
        if n < 100:
            return tens[n // 10] + ('-' + ones[n % 10] if n % 10 else '')
        if n < 1000:
            return ones[n // 100] + ' hundred' + (
                ' and ' + _num_words(n % 100) if n % 100 else '')
        return str(n)

    pounds = int(price)
    pence = round((price - pounds) * 100)
    if pence >= 100:
        pounds += pence // 100
        pence %= 100
    pence = max(pence, 0)

    parts = []
    if pounds > 0:
        parts.append(_num_words(pounds) + " pound" + ("s" if pounds != 1 else ""))
    if pence > 0:
        parts.append(_num_words(pence))
    if not parts:
        return "zero pence"
    return " ".join(parts)


def format_date_for_speech(date_str: str) -> str:
    try:
        d = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday',
                'Saturday', 'Sunday']
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        day_num = d.day
        if 10 <= day_num % 100 <= 20:
            sfx = 'th'
        else:
            sfx = {1: 'st', 2: 'nd', 3: 'rd'}.get(day_num % 10, 'th')
        return f"{days[d.weekday()]} the {day_num}{sfx} of {months[d.month - 1]}"
    except Exception:
        return date_str


def format_time_for_speech(time_str: str) -> str:
    try:
        parts = time_str.split(":")
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        period = "in the morning" if hour < 12 else "in the afternoon"
        if hour == 0:
            hour = 12
            period = "in the morning"
        elif hour > 12:
            hour -= 12
        if minute == 0:
            return f"{hour} {period}"
        if minute == 30:
            return f"half past {hour} {period}"
        return f"{hour}:{minute:02d} {period}"
    except Exception:
        return time_str


def format_tyre_size_for_speech(size: str) -> str:
    result = ""
    for i, ch in enumerate(size.strip()):
        if ch == '/':
            result += " "
        elif ch == 'R' and i > 0 and size[i - 1].isdigit():
            result += " R "
        else:
            result += ch
    return result


_BRAND_MAP = {
    'MICHELIN': 'Michelin', 'PIRELLI': 'Pirelli', 'CONTINENTAL': 'Continental',
    'BRIDGESTONE': 'Bridgestone', 'GOODYEAR': 'Goodyear', 'DUNLOP': 'Dunlop',
    'YOKOHAMA': 'Yokohama', 'HANKOOK': 'Hankook', 'KUMHO': 'Kumho',
    'TOYO': 'Toyo', 'FALKEN': 'Falken', 'NEXEN': 'Nexen',
    'VREDESTEIN': 'Vredestein', 'UNIROYAL': 'Uniroyal', 'AVON': 'Avon',
    'RADAR': 'Radar', 'ZETA': 'Zeta', 'TRACMAX': 'Tracmax',
    'ACCELERA': 'Accelera', 'LASSA': 'Lassa', 'NOKIAN': 'Nokian',
    'MARSHAL': 'Marshal', 'OVATION': 'Ovation', 'WINRUN': 'Winrun',
}


def format_brand_for_speech(brand: str) -> str:
    return _BRAND_MAP.get(brand.upper().strip(), brand.title() if brand.isupper() else brand)


_MAKE_MAP = {
    'BMW': 'B M W', 'VW': 'V W', 'VOLKSWAGEN': 'Volkswagen',
    'MERCEDES': 'Mercedes', 'MERCEDES-BENZ': 'Mercedes Benz',
    'PEUGEOT': 'Per-zho', 'RENAULT': 'Renno', 'CITROEN': 'Citron',
    'FIAT': 'Fee-at', 'SEAT': 'Say-at', 'SKODA': 'Shkoda',
    'HYUNDAI': 'Hyun-day', 'LAND ROVER': 'Land Rover',
    'RANGE ROVER': 'Range Rover',
}


def _clean_model_name(model: str) -> str:
    """Strip trim/variant codes that TTS reads letter-by-letter.
    'RANGE ROVER EVOQUE R-DYN S D A' → 'Range Rover Evoque'
    'FIESTA ST-LINE X' → 'Fiesta'
    """
    # Split on common trim code patterns: single letters, dashes, codes like R-DYN, ST-LINE
    import re
    # Remove anything after a token that looks like a trim code
    # Trim codes: single uppercase letters, X-prefixed, hyphenated codes, all-caps short tokens
    words = model.strip().split()
    clean = []
    for w in words:
        # Stop at single-letter tokens (S, D, A, X) or hyphenated codes (R-DYN, ST-LINE)
        if len(w) == 1 and w.isalpha():
            break
        if '-' in w and len(w) <= 8:
            break
        clean.append(w.title())
    return " ".join(clean) if clean else model.title()


def format_vehicle_for_speech(make: str, model: str) -> str:
    fm = _MAKE_MAP.get(make.upper(), make)
    cleaned_model = _clean_model_name(model)
    # Avoid repeating make in model (e.g. "Land Rover Range Rover Evoque")
    fm_lower = fm.lower()
    cm_lower = cleaned_model.lower()
    if cm_lower.startswith(fm_lower):
        cleaned_model = cleaned_model[len(fm):].strip()
    return f"{fm} {cleaned_model}"


# ═══════════════════════════════════════════════════════════════════════════
# VRN NORMALIZATION — NATO decode, digit words, make/model strip, max 7
# ═══════════════════════════════════════════════════════════════════════════
def normalize_vrn(raw: str) -> str:
    """Normalize caller's spoken registration into uppercase alphanumeric.
    Handles NATO phonetics, digit words, car make/model words, and max 7 chars."""
    tokens = re.split(r"[\s,.\-]+", raw.strip().lower())
    result = []
    for tok in tokens:
        if not tok:
            continue
        if tok in _NATO_LETTER_MAP:
            result.append(_NATO_LETTER_MAP[tok])
        elif tok in _DIGIT_WORD_MAP:
            result.append(_DIGIT_WORD_MAP[tok])
        elif tok in _CAR_MAKE_MODEL_WORDS:
            continue  # strip car make/model words
        elif tok.startswith("double") and len(tok) > 6:
            letter = tok[6:]
            if letter in _NATO_LETTER_MAP:
                result.extend([_NATO_LETTER_MAP[letter]] * 2)
            elif len(letter) == 1 and letter.isalnum():
                result.extend([letter.upper()] * 2)
        else:
            # Take alphanumeric characters
            for ch in tok:
                if ch.isalnum():
                    result.append(ch.upper())
    normalized = "".join(result)[:7]  # UK VRNs max 7 characters
    return normalized


# ═══════════════════════════════════════════════════════════════════════════
# INPUT SANITISATION
# ═══════════════════════════════════════════════════════════════════════════
def sanitise_phone(raw: str) -> str:
    """Fix 'flush 44' → '+44', convert spoken words to digits, and clean."""
    cleaned = raw.strip().lower()
    for word in _PLUS_MISHEARINGS:
        if cleaned.startswith(word):  # NOT regex \b — digits fused to word
            cleaned = "+" + cleaned[len(word):]
            break
    # Convert spoken number words to digits BEFORE stripping non-digits
    # Handle "hundred" (e.g. "zero eight hundred" → "0800")
    cleaned = cleaned.replace("hundred", "00")
    cleaned = cleaned.replace("thousand", "000")
    # Replace spoken digit words with their digit equivalents
    for word, digit in _SPOKEN_DIGITS.items():
        cleaned = re.sub(r'\b' + word + r'\b', digit, cleaned)
    # Handle "double X" → "XX" (e.g. "double seven" → "77")
    cleaned = re.sub(r'\bdouble\s+(\d)', r'\1\1', cleaned)
    # Handle "triple X" → "XXX"
    cleaned = re.sub(r'\btriple\s+(\d)', r'\1\1\1', cleaned)
    digits = re.sub(r"[^0-9+]", "", cleaned)
    if "+" in digits:
        digits = "+" + digits.replace("+", "")
    return digits or raw


def sanitise_email(raw: str) -> str:
    """Fix spoken digits and collapse underscore-separated digits."""
    cleaned = raw.strip().lower().replace(" ", "")
    for word, digit in _SPOKEN_DIGITS.items():
        cleaned = cleaned.replace(word, digit)
    # Collapse _1_2_3 → _123
    while re.search(r"(\d)[_](\d)", cleaned):
        cleaned = re.sub(r"(\d)[_](\d)", r"\1\2", cleaned)
    return cleaned


def validate_postcode(raw: str) -> Optional[str]:
    """Validate UK postcode. Returns cleaned postcode or None."""
    cleaned = raw.strip().upper().replace(" ", "")
    if "@" in cleaned or "." in cleaned and len(cleaned) > 8:
        return None  # email
    if cleaned.startswith("+") or (cleaned.isdigit() and len(cleaned) > 6):
        return None  # phone number
    if cleaned.isalpha() or cleaned.isdigit():
        return None  # must have both letters and digits
    if len(cleaned) < 5 or len(cleaned) > 8:
        return None
    # Insert space before last 3 chars for standard format
    return cleaned[:-3] + " " + cleaned[-3:]


# ═══════════════════════════════════════════════════════════════════════════
# ERROR MONITORING — Discord + Excel
# ═══════════════════════════════════════════════════════════════════════════
DISCORD_COLORS = {
    "error": 0xFF0000, "warning": 0xFFA500, "success": 0x00FF00,
    "info": 0x0099FF, "test": 0x9B59B6,
}

EXCEL_HEADERS = [
    "Timestamp", "Error Type", "Endpoint", "Status Code", "Response Time",
    "Request URL", "Request Body", "Session ID", "Error Message",
    "Notes", "Fixed?", "Fixed By", "Fix Date",
]


async def send_discord_notification(
    title: str, description: str = "", color: str = "info",
    fields: list = None, is_test: bool = False,
):
    if not DISCORD_WEBHOOK_URL:
        return
    if is_test:
        title = f"[TEST] {title}"
        color = "test"
    embed = {
        "title": title, "description": description,
        "color": DISCORD_COLORS.get(color, DISCORD_COLORS["info"]),
        "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "footer": {"text": f"Tyresoft Agent | {TYRESOFT_WORKSPACE}"},
    }
    if fields:
        embed["fields"] = fields
    try:
        async with aiohttp.ClientSession() as sess:
            async with sess.post(
                DISCORD_WEBHOOK_URL, json={"embeds": [embed]},
                headers={"Content-Type": "application/json"},
            ) as resp:
                if resp.status == 204:
                    print(f"[DISCORD] Sent: {title}")
                else:
                    print(f"[DISCORD] Failed: {resp.status}")
    except Exception as e:
        print(f"[DISCORD] Error: {e}")


def log_error_to_excel(
    error_type: str, endpoint: str, status_code: str = "N/A",
    response_time: str = "N/A", request_url: str = "", request_body: str = "",
    session_id: str = "N/A", error_message: str = "",
):
    if not HAS_OPENPYXL:
        print(f"[EXCEL] openpyxl not installed, skipping log")
        return
    try:
        fp = Path(ERROR_LOG_PATH)
        if not fp.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Error Log"
            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for col, h in enumerate(EXCEL_HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = hf
                cell.fill = hfill
                cell.alignment = Alignment(horizontal='center')
            ws.freeze_panes = 'A2'
            wb.save(fp)
        wb = load_workbook(fp)
        ws = wb.active
        nxt = ws.max_row + 1
        row = [
            uk_timestamp(), error_type, endpoint, str(status_code), response_time,
            request_url, (request_body[:1000] if request_body else ""),
            session_id, (error_message[:500] if error_message else ""),
            "", "", "", "",
        ]
        efill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=nxt, column=col, value=val)
            if col <= 9:
                cell.fill = efill
        wb.save(fp)
        print(f"[EXCEL] Logged row {nxt}: {endpoint}")
    except Exception as e:
        print(f"[EXCEL] Error: {e}")


async def send_api_error_notification(
    error_type: str, endpoint: str, status_code: str = "N/A",
    response_time: str = "N/A", request_url: str = "", request_body: str = "",
    session_id: str = "N/A", error_message: str = "",
):
    log_error_to_excel(
        error_type=error_type, endpoint=endpoint, status_code=status_code,
        response_time=response_time, request_url=request_url,
        request_body=request_body, session_id=session_id, error_message=error_message,
    )
    fields = [
        {"name": "Error Type", "value": error_type, "inline": True},
        {"name": "Endpoint", "value": endpoint, "inline": True},
        {"name": "Status", "value": str(status_code), "inline": True},
        {"name": "Time", "value": response_time, "inline": True},
        {"name": "URL", "value": f"`{request_url}`", "inline": False},
    ]
    if request_body:
        preview = request_body[:800] + "..." if len(request_body) > 800 else request_body
        fields.append({"name": "Body", "value": f"```json\n{preview}\n```", "inline": False})
    fields.append({"name": "Error", "value": f"```{error_message[:500]}```" if error_message else "N/A", "inline": False})
    await send_discord_notification(
        title="API ERROR", description="Tyresoft API error.", color="error", fields=fields,
    )


# ═══════════════════════════════════════════════════════════════════════════
# TYRESOFT API CLIENT
# ═══════════════════════════════════════════════════════════════════════════
def _get_auth_headers() -> dict:
    creds = f"{TYRESOFT_API_USERNAME}:{TYRESOFT_API_PASSWORD}"
    encoded = base64.b64encode(creds.encode()).decode()
    return {
        "Authorization": f"Basic {encoded}",
        "x-api-key": TYRESOFT_X_API_KEY,
        "Content-Type": "application/json",
    }


async def _api_call(
    method: str, url: str, body: dict = None, endpoint_name: str = "",
    session_id: str = "N/A",
) -> Optional[Any]:
    """Common API call with error handling, Discord alerts, and Excel logging."""
    start = time.time()
    body_str = json.dumps(body, indent=2) if body else "N/A"
    try:
        async with aiohttp.ClientSession(timeout=API_TIMEOUT) as sess:
            if method == "GET":
                resp = await sess.get(url, headers=_get_auth_headers())
            else:
                resp = await sess.post(url, headers=_get_auth_headers(), json=body)
            elapsed = time.time() - start
            if elapsed > 5.0:
                print(f"[API] SLOW {endpoint_name}: {elapsed:.1f}s")
            if resp.status == 200:
                return await resp.json()
            error_text = await resp.text()
            print(f"[API] {endpoint_name} error {resp.status}: {error_text[:200]}")
            await send_api_error_notification(
                error_type="HTTP Error", endpoint=endpoint_name,
                status_code=resp.status, response_time=f"{elapsed:.2f}s",
                request_url=url, request_body=body_str if method != "GET" else "GET",
                session_id=session_id, error_message=error_text,
            )
    except asyncio.TimeoutError:
        elapsed = time.time() - start
        print(f"[API] {endpoint_name} timeout")
        await send_api_error_notification(
            error_type="TimeoutError", endpoint=endpoint_name,
            response_time=f"{elapsed:.2f}s", request_url=url,
            request_body=body_str if method != "GET" else "GET",
            session_id=session_id, error_message="Request timed out (30s)",
        )
    except Exception as e:
        elapsed = time.time() - start
        print(f"[API] {endpoint_name} exception: {e}")
        await send_api_error_notification(
            error_type=type(e).__name__, endpoint=endpoint_name,
            response_time=f"{elapsed:.2f}s", request_url=url,
            request_body=body_str if method != "GET" else "GET",
            session_id=session_id, error_message=str(e),
        )
    return None


async def lookup_vehicle_by_vrm(vrm: str) -> Optional[dict]:
    url = f"{TYRESOFT_BASE_URL}/vrmLookup/{vrm}"
    print(f"[API] VRM Lookup: {url}")
    return await _api_call("GET", url, endpoint_name="vrmLookup")


async def get_available_slots(depot_id: int, service_ids: list, start_date: str) -> list:
    url = f"{TYRESOFT_BASE_URL}/availableSlotsForBasket/{depot_id}/{start_date}"
    body = {"list": service_ids if service_ids else [0]}
    result = await _api_call("POST", url, body=body, endpoint_name="availableSlotsForBasket")
    return result if isinstance(result, list) else []


async def save_customer(
    first_name: str, last_name: str = "", mobile: str = "", email: str = "",
) -> Optional[dict]:
    url = f"{TYRESOFT_BASE_URL}/saveCustomer"
    body = {
        "customerID": 0, "accountNumber": "",
        "contactData": {
            "name": {"salutation": "", "firstName": first_name, "lastName": last_name, "company": ""},
            "address": {
                "addressLine1": "", "addressLine2": "", "addressLine3": "", "addressLine4": "",
                "city": "", "county": "", "postcode": "", "country": "", "longitude": "", "latitude": "",
            },
            "contact": {"contact": "", "mobile": mobile, "email": email, "telephone": "", "twitter": ""},
            "sendSMSCorrespondance": False, "sendEmailCorrespondance": False,
            "sendPostalCorrespondance": False, "marketingOptOut": False,
        },
        "priceLevelID": 0, "creditAccount": False, "notes": "",
    }
    print(f"[API] Saving customer: {first_name} {last_name}")
    return await _api_call("POST", url, body=body, endpoint_name="saveCustomer")


async def save_vehicle(vrm: str, make: str, model: str, vehicle_info: dict = None) -> Optional[dict]:
    url = f"{TYRESOFT_BASE_URL}/saveVehicle"
    vi = vehicle_info or {}
    body = {
        "vehicleID": 0,
        "specifications": {
            "vrm": vrm, "make": make, "model": model,
            "yearOfManufacture": vi.get("yearOfManufacture", ""),
            "colour": vi.get("colour", ""),
            "mvrisMakeCode": "", "mvrisModelCode": "",
            "vinSerialNo": vi.get("vinSerialNo", ""),
            "dateFirstRegistered": vi.get("dateFirstRegistered", ""),
            "engineCapacity": vi.get("engineCapacity", ""),
            "transmission": vi.get("transmission", ""),
            "fuel": vi.get("fuel", ""),
            "doorplan": vi.get("doorplan", ""),
            "engineNumber": "", "co2Emissions": "", "gears": "",
            "motDue": vi.get("motDue", ""),
            "taxDue": "", "lastVRMLookupDate": "",
            "tyreSizeOptions": [],
        },
        "tyreSize": {
            "tyreSizeFront": "", "speedRatingFront": "", "loadIndexFront": "",
            "tyrePressureFront": "", "tyreSizeRear": "", "speedRatingRear": "",
            "loadIndexRear": "", "tyrePressureRear": "",
        },
        "customerID": 0, "motDueDate": "", "taxDueDate": "", "serviceDueDate": "",
        "tyreCheckDate": "", "nextInspectionDate": "", "authorisedVehicle": False,
        "fleetNumber": "", "vrmChecked": False,
        "flagData": {"flagName": "", "flagNotes": ""},
    }
    print(f"[API] Saving vehicle: {vrm}")
    return await _api_call("POST", url, body=body, endpoint_name="saveVehicle")


async def create_sale(
    depot_id: int, customer_id: int, vehicle_id: int,
    booking_slot: dict, items: list,
) -> Optional[dict]:
    url = f"{TYRESOFT_BASE_URL}/createSale"
    body = {
        "depotID": depot_id,
        "saleDate": booking_slot.get("date", uk_date()),
        "saleStatus": "Order",
        "notes": "Booking created via Reception Mate Voice AI",
        "worksheetNumber": "", "salesAdvisorID": 0,
        "poNumber": f"RM-{int(time.time())}",
        "flag": 1, "flagNotes": "Reception Mate Booking",
        "advertisingSurvey": "",
        "customerID": customer_id,
        "currencyUnit": {"currencyCode": "", "conversionRate": 0},
        "vehicleID": vehicle_id,
        "vehicleMileage": 0,
        "channelID": CHANNEL_ID,
        "orderStatus": "Awaiting Acknowledgement",
        "externalOrderReference": "", "channelBuyer": "",
        "overrideInvoiceNumber": "",
        "deliveryAddressID": 0, "deliveryType": "NONE",
        "sourceShippingOverride": "",
        "fittingCentreID": 0, "deliverToFittingCentre": False,
        "workSummary": "", "advisoryNotes": "",
        "bookingSlot": {
            "date": booking_slot.get("date", ""),
            "time": booking_slot.get("time", ""),
            "diaryCategoryID": booking_slot.get("diaryCategoryID", 1),
            "estimatedTime": booking_slot.get("estimatedTime", 30),
            "slotTypeID": booking_slot.get("slotTypeID", 1),
        },
        "items": items,
        "holdUntilDate": "", "authorisePayment": "",
        "payments": [{
            "paymentMethodID": 0, "paymentAmount": 0, "paymentDate": "",
            "paymentReference": "", "externalReference": "",
            "leaveUnallocated": True, "depotID": 0,
            "overrideDepositAccountID": 0, "customerID": 0,
        }],
        "customGroupID": 0, "customValues": [],
        "vatOverrideAmount": 0, "grossTotalForVATOverride": 0,
        "gsQuoteJobNumber": 0, "collectionSourceSaleLineID": 0,
    }
    print(f"[API] Creating sale: depot={depot_id} customer={customer_id} vehicle={vehicle_id}")
    print(f"[API] Sale items ({len(items)}):")
    for i, itm in enumerate(items, 1):
        svc_id = itm.get("serviceID", 0)
        code = itm.get("itemCode", "")
        qty = itm.get("quantity", 1)
        cost = itm.get("unitCost", 0)
        kind = "SERVICE" if svc_id else "TYRE"
        print(f"  [{i}] {kind}: code={code!r} svcID={svc_id} qty={qty} unit={cost}")
    return await _api_call("POST", url, body=body, endpoint_name="createSale")


def build_tyre_item(stock_number: str, quantity: int, unit_price: float) -> dict:
    return {
        "saleLineID": 0, "productID": 0, "tyrecatID": 0,
        "productEANCode": "", "productManufacturerCode": "",
        "serviceID": 0,  # 0 = tyre item
        "shippingService": False, "incomeAccountID": 0, "sequence": 0,
        "itemCode": stock_number,
        "itemDescription": "", "recordedDescription": "",
        "technicianID": 0,
        "quantity": quantity,
        "unitCost": unit_price,
        "unitCostIncludesVAT": False,
        "discount": 0, "vatCodeID": 0, "backOrderQuantity": 0,
        "taggedItemIdentifier": "", "linkLineID": 0,
        "hideChildLinks": False, "groupLinkSellPrices": False,
        "voucherCode": "", "voucherCodeLine": False,
        "estimatedCost": 0, "protectEstimatedCost": False,
        "leadTime": 0, "sourceSupplierID": 0, "sourcePurchaseOrderID": 0,
        "externalOrderLineReference": "",
        "changeInQtyAffectingPickList": False, "creditedAmount": 0,
    }


def build_service_item(service_id: int, unit_price: float) -> dict:
    return {
        "saleLineID": 0, "productID": 0, "tyrecatID": 0,
        "productEANCode": "", "productManufacturerCode": "",
        "serviceID": service_id,
        "shippingService": False, "incomeAccountID": 0, "sequence": 0,
        "itemCode": "",
        "itemDescription": "", "recordedDescription": "",
        "technicianID": 0,
        "quantity": 1,
        "unitCost": unit_price,
        "unitCostIncludesVAT": False,
        "discount": 0, "vatCodeID": 0, "backOrderQuantity": 0,
        "taggedItemIdentifier": "", "linkLineID": 0,
        "hideChildLinks": False, "groupLinkSellPrices": False,
        "voucherCode": "", "voucherCodeLine": False,
        "estimatedCost": 0, "protectEstimatedCost": False,
        "leadTime": 0, "sourceSupplierID": 0, "sourcePurchaseOrderID": 0,
        "externalOrderLineReference": "",
        "changeInQtyAffectingPickList": False, "creditedAmount": 0,
    }


# ═══════════════════════════════════════════════════════════════════════════
# TYRE INVENTORY (CSV)
# ═══════════════════════════════════════════════════════════════════════════
TYRE_INVENTORY: dict[int, list[dict]] = {}


def load_tyre_inventory():
    global TYRE_INVENTORY
    for branch_id in [1, 2]:
        TYRE_INVENTORY[branch_id] = []
        filename = f"Products Branch {branch_id}.csv"
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    TYRE_INVENTORY[branch_id].append({
                        "stock_number": row.get("Product Stock Number", ""),
                        "ean": row.get("Product EAN", ""),
                        "title": row.get("Product Title", ""),
                        "price": float(row.get("Retail", 0) or 0),
                        "width": row.get("Width", ""),
                        "aspect_ratio": row.get("Aspect Ratio", ""),
                        "rim": row.get("Rim", ""),
                        "speed_rating": row.get("Speed Rating", ""),
                        "load_index": row.get("Load Index", ""),
                        "brand": row.get("Brand Name", ""),
                        "vehicle_type": row.get("Vehicle Type", ""),
                        "product_type": row.get("Product Type", ""),
                        "runflat": row.get("Runflat", "FALSE").upper() == "TRUE",
                        "availability": row.get("Product Channel Available", ""),
                        "lead_time": row.get("Product Channel Lead Time", ""),
                    })
            print(f"[STARTUP] Loaded {len(TYRE_INVENTORY[branch_id])} tyres for Branch {branch_id}")
        except FileNotFoundError:
            print(f"[STARTUP] Warning: {filename} not found")
        except Exception as e:
            print(f"[STARTUP] Error loading {filename}: {e}")


def search_inventory(
    branch_id: int, width: str = "", aspect: str = "", rim: str = "",
    speed_rating: str = "", brand: str = "", max_results: int = 6,
) -> list[dict]:
    """Search tyre inventory by size components and optional brand."""
    matches = []
    for tyre in TYRE_INVENTORY.get(branch_id, []):
        if width and tyre.get("width") != width:
            continue
        if aspect and tyre.get("aspect_ratio") != aspect:
            continue
        if rim and tyre.get("rim") != rim:
            continue
        if speed_rating and tyre.get("speed_rating", "").upper() != speed_rating.upper():
            continue
        if brand and brand.upper() not in tyre.get("brand", "").upper():
            continue
        matches.append(tyre)
    matches.sort(key=lambda x: x.get("price", 999999))
    return matches[:max_results]


def parse_tyre_size(size_str: str) -> dict:
    """Parse '205/55R16 91V' or '235/60R18 107V' or '205 55 16' into width, aspect, rim, speed.
    Format: {width}/{aspect}R{rim} {load_index}{speed_rating}
    The space before load index is critical — don't strip it."""
    clean = size_str.upper().strip()
    result = {"width": "", "aspect": "", "rim": "", "speed": ""}
    if "/" in clean:
        # Split on space FIRST to separate size from load/speed
        # "235/60R18 107V" → ["235/60R18", "107V"]
        space_parts = clean.split()
        size_part = space_parts[0]  # "235/60R18"
        load_speed = space_parts[1] if len(space_parts) > 1 else ""  # "107V"

        parts = size_part.split("/")
        result["width"] = parts[0]
        remaining = parts[1] if len(parts) > 1 else ""
        if "R" in remaining:
            rp = remaining.split("R")
            result["aspect"] = rp[0]
            result["rim"] = rp[1] if len(rp) > 1 else ""

        # Extract speed rating from load/speed part (e.g. "107V" → speed "V")
        if load_speed:
            result["speed"] = load_speed[-1] if load_speed[-1].isalpha() else ""
    else:
        no_space = clean.replace(" ", "")
        if len(no_space) >= 7 and no_space[:3].isdigit():
            result["width"] = no_space[:3]
            result["aspect"] = no_space[3:5]
            result["rim"] = no_space[5:7]
            if len(no_space) > 7 and no_space[7].isalpha():
                result["speed"] = no_space[7]
    return result


# ═══════════════════════════════════════════════════════════════════════════
# SPECIALIST LLMs — LiveKit Cloud Inference Gateway
# ═══════════════════════════════════════════════════════════════════════════
def _create_inference_token() -> str:
    """Mint a short-lived JWT for LiveKit Cloud inference gateway."""
    try:
        from livekit import api as lk_api
        grant = lk_api.access_token.InferenceGrants(perform=True)
        return (
            lk_api.AccessToken(
                os.getenv("LIVEKIT_API_KEY", ""),
                os.getenv("LIVEKIT_API_SECRET", ""),
            )
            .with_identity("agent")
            .with_inference_grants(grant)
            .with_ttl(datetime.timedelta(seconds=600))
            .to_jwt()
        )
    except Exception as e:
        print(f"[SPECIALIST] Failed to create inference token: {e}")
        return ""


def _get_specialist_llm() -> Optional[Any]:
    """Lazy-init AsyncOpenAI client routed through LiveKit Cloud inference."""
    global _specialist_llm
    if not HAS_OPENAI:
        return None
    if _specialist_llm is None:
        if not os.getenv("LIVEKIT_API_KEY") or not os.getenv("LIVEKIT_API_SECRET"):
            return None
        token = _create_inference_token()
        if not token:
            return None
        _specialist_llm = AsyncOpenAI(
            api_key=token, base_url=_LIVEKIT_INFERENCE_URL,
        )
    else:
        # Refresh JWT before each use (10-min TTL)
        token = _create_inference_token()
        if token:
            _specialist_llm.api_key = token
    return _specialist_llm


# --- Service Advisor (temp 0.1) ---
_SERVICE_ADVISOR_PROMPT = """You are an automotive service advisor at a UK tyre centre.
Given the customer's description and available services, pick the single most suitable option.

Available services:
{service_list}

Rules:
- "tyres worn", "need new tyres", "need rubbers", "bald tyres" → "need_tyres"
- "steering pulls", "tracking", "alignment", "car drifts" → "WA"
- "MOT", "MOT test", "MOT due", "MOT expired" → "MOT-4"
- "full service", "service overdue", "hasn't been serviced in ages" → "full_service" (needs engine size)
- "air con", "AC not cold", "climate control" → "AIR1"
- "flat tyre", "puncture", "nail in tyre", "slow puncture" → "PUNC"
- "hybrid service" → "FSE1"
- Strange noises, warning lights, unknown issues → "callback" (can't diagnose over phone)

Reply with JSON ONLY:
{{"service_code": "code or need_tyres or full_service or callback", "reason": "one sentence", "needs_vrm": true or false}}"""


async def ask_service_advisor(description: str) -> dict:
    """Ask specialist to match a vague description to a service."""
    svc_list = "\n".join(f"- {k}: {v['name']} ({v['price']} pounds)" for k, v in SERVICES.items())
    prompt = _SERVICE_ADVISOR_PROMPT.format(service_list=svc_list)
    client = _get_specialist_llm()
    if client:
        try:
            resp = await asyncio.wait_for(
                client.chat.completions.create(
                    model=SPECIALIST_MODEL,
                    messages=[
                        {"role": "system", "content": prompt},
                        {"role": "user", "content": description},
                    ],
                    temperature=0.1,
                ),
                timeout=SPECIALIST_TIMEOUT,
            )
            raw = resp.choices[0].message.content.strip()
            raw = re.sub(r"^```json\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            return json.loads(raw)
        except Exception as e:
            print(f"[SPECIALIST] Service advisor error: {e}")
    # Fallback: keyword matching
    return _fuzzy_service_match(description)


def _fuzzy_service_match(description: str) -> dict:
    d = description.lower()
    if any(w in d for w in ("tyre", "tire", "rubber", "bald", "worn tyre")):
        return {"service_code": "need_tyres", "reason": "Customer needs tyres", "needs_vrm": True}
    if any(w in d for w in ("align", "tracking", "pull", "drift")):
        return {"service_code": "WA", "reason": "Alignment issue", "needs_vrm": False}
    if "mot" in d:
        return {"service_code": "MOT-4", "reason": "MOT test", "needs_vrm": False}
    if any(w in d for w in ("service", "overdue", "hasn't been")):
        return {"service_code": "full_service", "reason": "Service needed", "needs_vrm": True}
    if any(w in d for w in ("air con", "ac ", "climate")):
        return {"service_code": "AIR1", "reason": "Air con recharge", "needs_vrm": False}
    if any(w in d for w in ("puncture", "flat", "nail")):
        return {"service_code": "PUNC", "reason": "Puncture repair", "needs_vrm": False}
    if any(w in d for w in ("hybrid", "electric")):
        return {"service_code": "FSE1", "reason": "Hybrid service", "needs_vrm": False}
    return {"service_code": "callback", "reason": "Needs human assistance", "needs_vrm": False}


# --- Timeslot Matcher (temp 0) ---
_TIMESLOT_MATCHER_PROMPT = """You are a scheduling assistant at a UK tyre centre.
Given available booking slots and a customer's preference, find the best match.

Available slots (date and time):
{slots}

Today is: {today} ({today_name})
Tomorrow is: {tomorrow} ({tomorrow_name})

Rules:
- CRITICAL: If the caller asks for a SPECIFIC DAY (e.g. "Friday") and NO slot exists on that day, return the DATE for that day with time null. NEVER substitute a different day.
- "tomorrow" = {tomorrow}
- "by Friday", "before Friday", "by the end of Friday" = the ACTUAL day mentioned (Friday), NOT the day before. Return Friday's date.
- "next Monday/Tuesday/etc" or "Monday next week" = the FOLLOWING week's occurrence (7+ days from today), NOT this week
- IMPORTANT: "next week Thursday" or "Thursday next week" means the Thursday that falls in the NEXT calendar week, NOT the coming Thursday if one exists this week.
  For example if today is Wednesday {today}, "Thursday next week" = the Thursday 8 days away, NOT tomorrow.
- "morning" = before 12:00
- "afternoon" = 12:00-17:00
- "around 10" or "about 10" = closest to 10:00
- Prefer exact match, then closest time on same day
- CRITICAL: When returning a date, ALWAYS calculate it from today ({today} = {today_name}). Double-check your weekday arithmetic.
  Monday=0, Tuesday=1, Wednesday=2, Thursday=3, Friday=4, Saturday=5, Sunday=6.
- If the requested date has no available slots, return {{"date": "YYYY-MM-DD", "time": null}} so the system can fetch availability for that date.

Reply with JSON ONLY:
{{"date": "YYYY-MM-DD", "time": "HH:MM"}} or {{"date": null, "time": null, "reason": "why no match"}}"""


async def ask_timeslot_matcher(preference: str, available_slots: list) -> dict:
    """Ask specialist to parse natural language date/time preference."""
    now = uk_now()
    today = now.date()
    tomorrow = today + datetime.timedelta(days=1)
    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    # Format slots for prompt
    slot_lines = []
    for s in available_slots[:30]:  # limit to avoid huge prompts
        slot_lines.append(f"  {s.get('date', '')} {s.get('time', '')}")
    slots_text = "\n".join(slot_lines) if slot_lines else "No slots loaded"

    prompt = _TIMESLOT_MATCHER_PROMPT.format(
        slots=slots_text,
        today=today.isoformat(), today_name=day_names[today.weekday()],
        tomorrow=tomorrow.isoformat(), tomorrow_name=day_names[tomorrow.weekday()],
    )

    client = _get_specialist_llm()
    if client:
        try:
            resp = await asyncio.wait_for(
                client.chat.completions.create(
                    model=SPECIALIST_MODEL,
                    messages=[
                        {"role": "system", "content": prompt},
                        {"role": "user", "content": preference},
                    ],
                    temperature=0,
                ),
                timeout=SPECIALIST_TIMEOUT,
            )
            raw = resp.choices[0].message.content.strip()
            raw = re.sub(r"^```json\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            return json.loads(raw)
        except Exception as e:
            print(f"[SPECIALIST] Timeslot matcher error: {e}")
    # Fallback: basic parsing
    return _parse_timeslot_fallback(preference, available_slots, today)


def _parse_timeslot_fallback(preference: str, slots: list, today: datetime.date) -> dict:
    pref = preference.lower().strip()
    # Strip "by" prefix — "by Friday" means ON Friday, not before it
    pref = re.sub(r"^by\s+", "", pref)
    target_date = None
    target_time = None

    # Date parsing
    next_week = "next week" in pref or "next " in pref  # "next Thursday", "Thursday next week"
    if "tomorrow" in pref:
        target_date = (today + datetime.timedelta(days=1)).isoformat()
    elif "today" in pref:
        target_date = today.isoformat()
    else:
        day_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        for i, dn in enumerate(day_names):
            if dn in pref:
                days_ahead = i - today.weekday()
                if days_ahead <= 0:
                    days_ahead += 7
                # "next week" or "next Thursday" → push to FOLLOWING week
                if next_week and days_ahead <= 7:
                    days_ahead += 7
                target_date = (today + datetime.timedelta(days=days_ahead)).isoformat()
                break

    # Time parsing
    time_match = re.search(r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)?", pref)
    if time_match:
        h = int(time_match.group(1))
        m = int(time_match.group(2) or 0)
        ampm = time_match.group(3)
        if ampm == "pm" and h < 12:
            h += 12
        elif ampm == "am" and h == 12:
            h = 0
        target_time = f"{h:02d}:{m:02d}"

    if "morning" in pref and not target_time:
        target_time = "09:00"
    elif "afternoon" in pref and not target_time:
        target_time = "13:00"

    # Match against available slots
    if not target_date and slots:
        target_date = slots[0].get("date")

    best = None
    for s in slots:
        if target_date and s.get("date") != target_date:
            continue
        if target_time:
            st = s.get("time", "")[:5]
            if st == target_time:
                return {"date": s["date"], "time": st}
            if best is None:
                best = s
        else:
            if best is None:
                best = s

    if best:
        return {"date": best["date"], "time": best.get("time", "")[:5]}
    # Return parsed date/time even if no slots matched — the caller
    # may just need the date so the API can fetch that day's slots
    return {"date": target_date, "time": target_time}


# --- Message Summariser (temp 0.2) ---
_MESSAGE_SUMMARISER_PROMPT = """You are a tyre centre receptionist summarising a callback message.
The caller wants someone to call them back. Summarise their message for the team.

Reply with JSON ONLY:
{{"category": "tyre_enquiry|service_enquiry|complaint|callback|urgent|other", "summary": "2-3 sentences", "action": "what the team should do"}}"""


async def ask_message_summariser(message: str, caller_name: str = "", phone: str = "") -> dict:
    context = f"Caller: {caller_name}\nPhone: {phone}\nMessage: {message}"
    client = _get_specialist_llm()
    if client:
        try:
            resp = await asyncio.wait_for(
                client.chat.completions.create(
                    model=SPECIALIST_MODEL,
                    messages=[
                        {"role": "system", "content": _MESSAGE_SUMMARISER_PROMPT},
                        {"role": "user", "content": context},
                    ],
                    temperature=0.2,
                ),
                timeout=SPECIALIST_TIMEOUT,
            )
            raw = resp.choices[0].message.content.strip()
            raw = re.sub(r"^```json\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            return json.loads(raw)
        except Exception as e:
            print(f"[SPECIALIST] Message summariser error: {e}")
    return {"category": "callback", "summary": message[:200], "action": "Call back customer"}


# ═══════════════════════════════════════════════════════════════════════════
# DYNAMIC GREETING
# ═══════════════════════════════════════════════════════════════════════════
def get_dynamic_greeting() -> str:
    hour = uk_now().hour
    if hour < 12:
        tod = "morning"
    elif hour < 17:
        tod = "afternoon"
    else:
        tod = "evening"
    return (
        f"Good {tod}, you're through to Leah at {AGENT_BRANCH_NAME}. "
        "I can help with tyres, MOTs, servicing, or book you in for a fitting. "
        "Who am I speaking to?"
    )


# ═══════════════════════════════════════════════════════════════════════════
# STARTUP
# ═══════════════════════════════════════════════════════════════════════════
print(f"[STARTUP] Tyresoft Agent infrastructure loaded")
print(f"[STARTUP] Workspace: {TYRESOFT_WORKSPACE}")
print(f"[STARTUP] Speaking model: {SPEAKING_MODEL}")
print(f"[STARTUP] Specialist model: {SPECIALIST_MODEL}")
load_tyre_inventory()
