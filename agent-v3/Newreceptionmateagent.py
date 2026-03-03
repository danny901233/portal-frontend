"""
SUPERVISOR RECEPTIONMATE — Single-Agent Architecture
One speaking agent (Leah) + state machine + worker tools.
Replaces multi-agent handoff system to eliminate mid-call speech ticks.
"""

import asyncio
import logging
import os
import re
import json
import time
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from enum import Enum
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import datetime as _dt

import aiohttp
from dotenv import load_dotenv
from openai import AsyncOpenAI

from livekit import api as lk_api
from livekit.agents import (
    Agent,
    AgentSession,
    ChatMessage,
    JobContext,
    RunContext,
    WorkerOptions,
    WorkerType,
    cli,
)
from livekit.agents.llm import function_tool
from livekit.plugins import deepgram, elevenlabs, silero
from livekit.plugins.turn_detector.multilingual import MultilingualModel

# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("supervisor-reception")

# ============================================================
# SPECIALIST LLM CLIENT (routes through LiveKit Cloud inference)
# Uses LIVEKIT_API_KEY + LIVEKIT_API_SECRET — no separate OPENAI_API_KEY needed.
# ============================================================

_LIVEKIT_INFERENCE_URL = "https://agent-gateway.livekit.cloud/v1"
_specialist_llm: Optional[AsyncOpenAI] = None


def _create_inference_token() -> str:
    """Mint a short-lived JWT for LiveKit Cloud inference gateway."""
    grant = lk_api.access_token.InferenceGrants(perform=True)
    return (
        lk_api.AccessToken(
            os.getenv("LIVEKIT_API_KEY", ""),
            os.getenv("LIVEKIT_API_SECRET", ""),
        )
        .with_identity("agent")
        .with_inference_grants(grant)
        .with_ttl(_dt.timedelta(seconds=600))
        .to_jwt()
    )


def _get_specialist_llm() -> Optional[AsyncOpenAI]:
    """Lazy-init an AsyncOpenAI client routed through LiveKit Cloud inference.
    Uses LIVEKIT_API_KEY + LIVEKIT_API_SECRET (no separate OPENAI_API_KEY needed).
    Returns None if LiveKit credentials are missing."""
    global _specialist_llm
    if _specialist_llm is None:
        if not os.getenv("LIVEKIT_API_KEY") or not os.getenv("LIVEKIT_API_SECRET"):
            return None
        _specialist_llm = AsyncOpenAI(
            api_key=_create_inference_token(),
            base_url=_LIVEKIT_INFERENCE_URL,
        )
    else:
        # Refresh the JWT before each use (10-min TTL)
        _specialist_llm.api_key = _create_inference_token()
    return _specialist_llm


# ============================================================
# ENVIRONMENT
# ============================================================

_DOTENV_CANDIDATES = [
    Path(__file__).parent / ".env",
    Path(__file__).parent / ".env.local",
    Path(__file__).parent.parent / ".env",
    Path(__file__).parent.parent / ".env.local",
]
_loaded_env_files: list[Path] = []
for candidate in _DOTENV_CANDIDATES:
    if candidate.exists():
        load_dotenv(dotenv_path=candidate, override=True)
        logger.info(f"[SETUP] Loaded environment from {candidate}")
        _loaded_env_files.append(candidate)

if not _loaded_env_files:
    logger.warning("[SETUP] No .env file found; relying on process env")

# ============================================================
# DYNAMODB CLIENT FOR PORTAL CONFIGURATION
# ============================================================

try:
    import boto3
    from botocore.exceptions import BotoCoreError, ClientError
except ImportError:
    boto3 = None
    BotoCoreError = ClientError = Exception

_dynamo_client = None


def _get_dynamo_client():
    """Initialize DynamoDB client for loading agent configuration."""
    global _dynamo_client
    if _dynamo_client is not None:
        return _dynamo_client
    
    if boto3 is None:
        logger.warning("[DYNAMO_INIT] boto3 not installed; configuration loading from DynamoDB unavailable")
        _dynamo_client = None
        return None
    
    region = os.getenv("AWS_REGION") or os.getenv("AWS_DEFAULT_REGION") or "eu-west-2"
    aws_access_key = os.getenv("AWS_ACCESS_KEY_ID")
    aws_secret_key = os.getenv("AWS_SECRET_ACCESS_KEY")
    aws_session_token = os.getenv("AWS_SESSION_TOKEN")
    
    try:
        if aws_access_key and not aws_secret_key:
            logger.warning("[DYNAMO_INIT] AWS_ACCESS_KEY_ID is set but AWS_SECRET_ACCESS_KEY is missing")
            return None
        if aws_access_key and aws_secret_key:
            _dynamo_client = boto3.client(
                "dynamodb",
                region_name=region,
                aws_access_key_id=aws_access_key,
                aws_secret_access_key=aws_secret_key,
                aws_session_token=aws_session_token,
            )
        else:
            _dynamo_client = boto3.client("dynamodb", region_name=region)
        logger.info(f"[DYNAMO_INIT] DynamoDB client initialized for region: {region}")
    except (BotoCoreError, ClientError) as error:
        logger.warning(f"[DYNAMO_INIT] Failed to create DynamoDB client: {error}")
        _dynamo_client = None
    return _dynamo_client


def _deserialize_dynamodb_value(attr_value):
    """Convert DynamoDB AttributeValue format to Python objects."""
    if isinstance(attr_value, dict):
        if "S" in attr_value:
            return attr_value["S"]
        elif "N" in attr_value:
            return float(attr_value["N"]) if "." in attr_value["N"] else int(attr_value["N"])
        elif "BOOL" in attr_value:
            return attr_value["BOOL"]
        elif "NULL" in attr_value:
            return None
        elif "M" in attr_value:
            return {k: _deserialize_dynamodb_value(v) for k, v in attr_value["M"].items()}
        elif "L" in attr_value:
            return [_deserialize_dynamodb_value(item) for item in attr_value["L"]]
    return attr_value


def load_agent_config(garage_id: str) -> tuple[dict, list]:
    """Load agent configuration and knowledge base from DynamoDB."""
    client = _get_dynamo_client()
    if not garage_id:
        logger.warning("[LOAD_CONFIG] No garage_id provided")
        return {}, []
    if client is None:
        logger.warning("[LOAD_CONFIG] DynamoDB client is None - AWS credentials may be missing")
        return {}, []
    
    logger.info(f"[LOAD_CONFIG] Loading configuration for garage_id: {garage_id}")
    try:
        response = client.get_item(
            TableName="AgentConfig",
            Key={"garageId": {"S": garage_id}},
        )
    except (BotoCoreError, ClientError) as error:
        logger.warning(f"[LOAD_CONFIG] Failed to load agent config: {error}")
        return {}, []

    item = response.get("Item")
    if not item:
        logger.warning(f"[LOAD_CONFIG] No configuration found in DynamoDB for garage_id: {garage_id}")
        return {}, []

    # Check if configuration is stored as Map (new format) or String (old format)
    config_attr = item.get("configuration", {})
    if "M" in config_attr:
        configuration = _deserialize_dynamodb_value(config_attr)
        logger.info(f"[LOAD_CONFIG] Loaded configuration as Map with {len(configuration)} keys")
    elif "S" in config_attr:
        configuration_raw = config_attr.get("S", "")
        try:
            configuration = json.loads(configuration_raw) if configuration_raw else {}
            logger.info(f"[LOAD_CONFIG] Loaded configuration from JSON string with {len(configuration)} keys")
        except json.JSONDecodeError:
            logger.error("[LOAD_CONFIG] Failed to parse configuration JSON")
            configuration = {}
    else:
        logger.warning("[LOAD_CONFIG] Configuration has unexpected format")
        configuration = {}

    # Check if knowledgeBase is stored as List (new format) or String (old format)
    kb_attr = item.get("knowledgeBase", {})
    if "L" in kb_attr:
        knowledge_base = _deserialize_dynamodb_value(kb_attr)
        logger.info(f"[LOAD_CONFIG] Loaded knowledge base as List with {len(knowledge_base)} items")
    elif "S" in kb_attr:
        knowledge_raw = kb_attr.get("S", "")
        try:
            knowledge_base = json.loads(knowledge_raw) if knowledge_raw else []
            logger.info(f"[LOAD_CONFIG] Loaded knowledge base from JSON string with {len(knowledge_base)} items")
        except json.JSONDecodeError:
            logger.error("[LOAD_CONFIG] Failed to parse knowledge base JSON")
            knowledge_base = []
    else:
        knowledge_base = []
    
    return configuration, knowledge_base


def _clamp(value: float, minimum: float, maximum: float) -> float:
    """Clamp a value between min and max."""
    return max(minimum, min(maximum, value))


def _apply_agent_configuration(configuration: dict, knowledge_base: list) -> str:
    """Apply loaded configuration from DynamoDB to global variables.
    Returns agent_mode ('assist' or 'automate')."""
    global GH_CUSTOMER_ID, GH_LOCATION_ID, GH_API_KEY
    global AGENT_BRANCH_NAME, AGENT_GREETING_LINE
    global ELEVEN_VOICE_ID, ELEVEN_STABILITY, ELEVEN_SIMILARITY, ELEVEN_STYLE
    global AGENT_KNOWLEDGE_BASE, AGENT_CONFIGURATION
    global GARAGE_HOURS, SERVICE_TYPE
    
    if not isinstance(configuration, dict):
        configuration = {}
    
    # Store raw configuration
    AGENT_CONFIGURATION = configuration
    
    # Apply branch name
    branch_name_value = configuration.get("branchName") or configuration.get("businessName")
    if isinstance(branch_name_value, str) and branch_name_value.strip():
        AGENT_BRANCH_NAME = branch_name_value.strip()
        logger.info(f"[APPLY_CONFIG] Branch name: {AGENT_BRANCH_NAME}")
    
    # Apply greeting line
    greeting_line = (configuration.get("greetingLine") or "").strip()
    if greeting_line:
        AGENT_GREETING_LINE = greeting_line
        logger.info(f"[APPLY_CONFIG] Greeting line: {greeting_line[:50]}...")
    
    # Apply GarageHive settings
    garage_hive_settings = configuration.get("garageHiveSettings") or {}
    logger.info(f"[APPLY_CONFIG] GarageHive settings: {garage_hive_settings}")
    
    # Support both 'customerId' and 'instanceUrl' keys
    instance_value = garage_hive_settings.get("customerId") or garage_hive_settings.get("instanceUrl")
    if instance_value and instance_value.strip():
        GH_CUSTOMER_ID = instance_value.strip()
        logger.info(f"[APPLY_CONFIG] Set GH_CUSTOMER_ID to: {GH_CUSTOMER_ID}")
    
    location_value = garage_hive_settings.get("locationId")
    if location_value and str(location_value).strip():
        GH_LOCATION_ID = str(location_value).strip()
        logger.info(f"[APPLY_CONFIG] Set GH_LOCATION_ID to: {GH_LOCATION_ID}")
    
    api_key_value = garage_hive_settings.get("apiKey")
    if api_key_value and str(api_key_value).strip():
        GH_API_KEY = str(api_key_value).strip()
        logger.info(f"[APPLY_CONFIG] Set GH_API_KEY (length: {len(GH_API_KEY)})")
    
    # Apply voice selection
    voice_name = configuration.get("voice")
    if voice_name and isinstance(voice_name, str):
        voice_mapping = {
            "tom": "Fahco4VZzobUeiPqni1S",
            "leah": "rfkTsdZrVWEVhDycUYn9",
            "sophie": "fq1SdXsX6OokE10pJ4Xw",
            "gemma": "IosqM5LMIzqPfT0efhhy",
            "isobel": "h8eW5xfRUGVJrZhAFxqK",
            "fraser": "v2zbX16tJNtRIx8rSHDM",
        }
        normalized_voice = voice_name.strip().lower()
        if normalized_voice in voice_mapping:
            ELEVEN_VOICE_ID = voice_mapping[normalized_voice]
            logger.info(f"[APPLY_CONFIG] Voice set to: {voice_name} (ID: {ELEVEN_VOICE_ID})")
    
    # Apply voice settings overrides
    voice_settings = configuration.get("voiceSettings")
    if isinstance(voice_settings, dict):
        if "voiceId" in voice_settings and isinstance(voice_settings["voiceId"], str):
            voice_id_override = voice_settings["voiceId"].strip()
            if voice_id_override:
                ELEVEN_VOICE_ID = voice_id_override
                logger.info(f"[APPLY_CONFIG] Voice ID override: {ELEVEN_VOICE_ID}")
        if "elevenStability" in voice_settings:
            ELEVEN_STABILITY = float(voice_settings["elevenStability"])
        if "elevenSimilarity" in voice_settings:
            ELEVEN_SIMILARITY = float(voice_settings["elevenSimilarity"])
        if "elevenStyle" in voice_settings:
            ELEVEN_STYLE = float(voice_settings["elevenStyle"])
        logger.info(
            f"[APPLY_CONFIG] Eleven settings: voice_id={ELEVEN_VOICE_ID}, "
            f"stability={ELEVEN_STABILITY}, similarity={ELEVEN_SIMILARITY}, style={ELEVEN_STYLE}"
        )
    
    # Apply garage hours
    garage_hours_config = configuration.get("weeklyOpeningHours") or configuration.get("garageHours") or {}
    if isinstance(garage_hours_config, dict) and garage_hours_config:
        GARAGE_HOURS = garage_hours_config
        logger.info(f"[APPLY_CONFIG] Garage hours: {garage_hours_config}")
    
    # Apply service type (fast-fit vs full-service)
    allow_fast_fit_only = configuration.get("allowFastFitOnly")
    if allow_fast_fit_only is True:
        SERVICE_TYPE = "fast-fit"
        logger.info(f"[APPLY_CONFIG] Service type: fast-fit (allowFastFitOnly=True)")
    else:
        SERVICE_TYPE = "full-service"
        logger.info(f"[APPLY_CONFIG] Service type: full-service")
    
    # Apply knowledge base (trimmed to reasonable size)
    if isinstance(knowledge_base, list):
        max_docs = 5  # Limit number of documents
        max_chars = 6000  # Limit total characters
        
        trimmed_kb: list[dict] = []
        total_chars = 0
        for doc in knowledge_base:
            if len(trimmed_kb) >= max_docs:
                break
            if not isinstance(doc, dict):
                continue
            content = (doc.get("content") or "").strip()
            if not content:
                continue
            remaining = max_chars - total_chars
            if remaining <= 0:
                break
            if len(content) > remaining:
                content = content[:remaining].rstrip()
            trimmed_kb.append({
                "title": doc.get("title") or "Untitled",
                "content": content
            })
            total_chars += len(content)
        
        AGENT_KNOWLEDGE_BASE = trimmed_kb
        if trimmed_kb:
            logger.info(f"[APPLY_CONFIG] Loaded {len(trimmed_kb)} knowledge base documents ({total_chars} chars)")
        else:
            logger.info("[APPLY_CONFIG] Knowledge base is empty")
    else:
        AGENT_KNOWLEDGE_BASE = []
    
    # Read agentType from configuration
    agent_type = configuration.get("agentType", "automate")
    if isinstance(agent_type, str):
        agent_mode = agent_type.strip().lower()
    else:
        agent_mode = "automate"
    
    logger.info(f"[APPLY_CONFIG] Agent mode: {agent_mode}")
    return agent_mode


def refresh_agent_configuration(garage_id: str) -> str:
    """Load and apply configuration from DynamoDB.
    Returns agent_mode ('assist' or 'automate')."""
    configuration, knowledge_base = load_agent_config(garage_id)
    if not configuration:
        logger.warning(f"[LOAD_CONFIG] No configuration found for garage {garage_id}; using .env defaults")
        return "automate"
    logger.info(f"[LOAD_CONFIG] Loaded configuration for garage {garage_id}")
    agent_mode = _apply_agent_configuration(configuration, knowledge_base)
    return agent_mode


# ============================================================
# CONFIGURATION
# ============================================================

# Read PORTAL_GARAGE_ID early so we can load config from DynamoDB
PORTAL_GARAGE_ID = os.getenv("PORTAL_GARAGE_ID", "").strip()

# Global storage for configuration and knowledge base
AGENT_CONFIGURATION: dict = {}
AGENT_KNOWLEDGE_BASE: list = []
GARAGE_HOURS: dict = {}  # Garage opening hours
SERVICE_TYPE: str = "full-service"  # Service type: "fast-fit" or "full-service"

# Load configuration if PORTAL_GARAGE_ID is set
if PORTAL_GARAGE_ID:
    try:
        refresh_agent_configuration(PORTAL_GARAGE_ID)
        logger.info(f"[MODULE_INIT] Loaded configuration for garage: {PORTAL_GARAGE_ID}")
    except Exception as e:
        logger.error(f"[MODULE_INIT] Failed to load configuration: {e}")
        logger.info("[MODULE_INIT] Falling back to .env configuration")
else:
    logger.info("[MODULE_INIT] No PORTAL_GARAGE_ID set, using .env configuration")

# Read environment variable defaults (used as fallback if DynamoDB not configured)
DEFAULT_GH_CUSTOMER_ID = "devbc24_mpu"
DEFAULT_GH_LOCATION_ID = "399"
_GH_CUSTOMER_PLACEHOLDERS = {"", "your-customer-id", "your_customer_id", "customer-id"}
_GH_LOCATION_PLACEHOLDERS = {"", "your-location-id", "your_location_id", "your-location-code"}


def _resolve_env_value(raw_value: Optional[str], default: str, placeholders: set[str], var_name: str) -> str:
    cleaned = (raw_value or "").strip()
    if cleaned.lower() in placeholders:
        logger.warning(f"[SETUP] {var_name} not configured; defaulting to {default}")
        return default
    if not cleaned:
        logger.warning(f"[SETUP] {var_name} missing; defaulting to {default}")
        return default
    return cleaned


GH_CUSTOMER_ID = _resolve_env_value(os.getenv("GH_CUSTOMER_ID"), DEFAULT_GH_CUSTOMER_ID, _GH_CUSTOMER_PLACEHOLDERS, "GH_CUSTOMER_ID")
GH_LOCATION_ID = _resolve_env_value(os.getenv("GH_LOCATION_ID"), DEFAULT_GH_LOCATION_ID, _GH_LOCATION_PLACEHOLDERS, "GH_LOCATION_ID")
GH_API_KEY = os.getenv("GH_API_KEY", "")
AGENT_BRANCH_NAME = os.getenv("AGENT_BRANCH_NAME", "the garage")
AGENT_GREETING_LINE = os.getenv("AGENT_GREETING_LINE", "")
# ElevenLabs TTS — boss prefers ElevenLabs voice quality over Cartesia
# Turbo v2.5 is their fastest model (~150-300ms TTFA)
ELEVEN_VOICE_ID = os.getenv("ELEVEN_VOICE_ID", "21m00Tcm4TlvDq8ikWAM")
ELEVEN_TTS_MODEL = os.getenv("ELEVEN_TTS_MODEL", "eleven_turbo_v2_5")
ELEVEN_STABILITY = float(os.getenv("ELEVEN_STABILITY", "0.65"))
ELEVEN_SIMILARITY = float(os.getenv("ELEVEN_SIMILARITY", "0.75"))
ELEVEN_STYLE = float(os.getenv("ELEVEN_STYLE", "0.1"))

DISCORD_WEBHOOK_URL = os.getenv(
    "DISCORD_WEBHOOK_URL",
    "https://discord.com/api/webhooks/1470410320949285017/DIavHqUlXs2UPMossJ6418TA9ipTtplv6tBTYJKX9kSxSqF9wcHXQjVYeIkxhiLiUizM",
)
ERROR_LOG_EXCEL_PATH = Path(
    os.getenv("ERROR_LOG_EXCEL_PATH", str(Path(__file__).parent / "error_log.xlsx"))
)

# Portal API configuration
PORTAL_API_URL = os.getenv("PORTAL_API_URL", "https://portal.receptionmate.co.uk/api/calls")
PORTAL_WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET", "optional-shared-secret")
RECORDING_BASE_URL = os.getenv("RECORDING_BASE_URL", "").strip()  # e.g. https://storage.../recordings

# ============================================================
# PORTAL CALL LOGGING
# ============================================================

async def generate_call_summary(transcript: list, state) -> str:
    """Use GPT via LiveKit inference to generate a detailed structured call summary."""
    # Build plain-text transcript from entries
    lines = []
    for entry in transcript:
        speaker = entry.get("speaker", "unknown").capitalize()
        text = entry.get("text", "")
        if text:
            lines.append(f"{speaker}: {text}")
    transcript_text = "\n".join(lines)

    if not transcript_text or len(lines) < 2:
        return "No conversation was had."

    try:
        llm = _get_specialist_llm()
        if not llm:
            raise ValueError("No LLM client available")

        # Build context from state for the prompt
        customer_name = f"{state.customer_name_first} {state.customer_name_last}".strip() or "Customer"
        vehicle_info = state.vrn or "their vehicle"
        if state.vrn and state.vehicle_make and state.vehicle_model:
            vehicle_info = f"{state.vrn} ({state.vehicle_make} {state.vehicle_model})"

        booking_status = ""
        if state.step == Step.CONFIRMED:
            booking_status = "BOOKING WAS SUCCESSFULLY SUBMITTED TO THE SYSTEM."
        elif state.booking_date and state.booking_time:
            booking_status = "TIMESLOT WAS DISCUSSED BUT BOOKING WAS NOT COMPLETED. Customer needs a callback to finalize the booking."
        elif state.intent in ("message", "quote"):
            booking_status = "Customer requested a MESSAGE/CALLBACK."

        prompt = f"""Create a structured call summary from this garage receptionist call transcript.

BOOKING STATUS: {booking_status}

CRITICAL RULES:
1. ONLY summarize information that is ACTUALLY in the transcript below
2. DO NOT make up or infer details that aren't explicitly stated
3. Use the BOOKING STATUS above to determine if booking was confirmed or callback needed
4. If BOOKING STATUS says "NOT COMPLETED", you MUST state "Callback required to complete the booking"
5. If the customer only said "hello" then hung up, respond ONLY with "No conversation was had."

REQUIRED FORMAT:
Start with: "{customer_name} called regarding {vehicle_info}."

Then provide 2-4 paragraphs covering ONLY what was actually discussed:

Paragraph 1 - Purpose & Vehicle:
- Why they called (service needed, enquiry, complaint, etc.)
- Vehicle details if mentioned (make, model, registration, issues)

Paragraph 2 - Services/Booking Details:
- Specific work discussed, prices quoted
- If booking confirmed: "Booking confirmed for [DATE] at [TIME]"
- If not completed: "Booking was NOT finalized."

Paragraph 3 - Callback/Follow-up:
- If callback needed: "Callback required - [specific reason]"
- Only state "No callback required" if booking was SUCCESSFULLY SUBMITTED

Transcript:
{transcript_text}

Call Summary:"""

        response = await asyncio.wait_for(
            llm.chat.completions.create(
                model="openai/gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You write factual, detailed summaries of garage phone calls based ONLY on what was actually said in the transcript. Never invent names, registrations, or details not present in the transcript."},
                    {"role": "user", "content": prompt},
                ],
                max_tokens=500,
                temperature=0.1,
            ),
            timeout=10.0,
        )
        summary = response.choices[0].message.content.strip()
        if not summary or len(summary) < 20:
            return "No conversation was had."
        logger.info(f"[PORTAL] GPT summary generated ({len(summary)} chars)")
        return summary
    except Exception as e:
        logger.warning(f"[PORTAL] GPT summary failed, using fallback: {e}")
        # Fallback: build summary from state fields
        parts = []
        name = f"{state.customer_name_first} {state.customer_name_last}".strip() or "Customer"
        parts.append(f"{name} called")
        if state.vrn:
            parts.append(f"regarding {state.vrn}")
            if state.vehicle_make and state.vehicle_model:
                parts.append(f"({state.vehicle_make} {state.vehicle_model})")
        if state.booking_date:
            parts.append(f"Booking: {state.booking_date} at {state.booking_time}")
        elif state.message:
            parts.append(f"Message: {state.message}")
        elif state.intent:
            parts.append(f"Intent: {state.intent}")
        return ". ".join(parts) + "."


async def log_call_to_portal(
    garage_id: str,
    room_name: str,
    duration_seconds: int,
    transcript: list,
    summary: str,
    customer_name: str = "",
    customer_phone: str = "",
    registration_number: str = "",
    confirmed_booking: bool = False,
    booking_details: str = "",
    call_type: str = "unknown",
    metrics: dict = None,
) -> None:
    """Log call data to the portal backend."""
    
    # Only log calls that are 55 seconds or longer
    if duration_seconds < 55:
        logger.info(f"[PORTAL] Skipping call log - duration {duration_seconds}s is under 55s threshold")
        return
    
    try:
        # metrics must be non-empty object
        if not metrics:
            metrics = {"duration_seconds": duration_seconds, "call_type": call_type}

        payload = {
            "garageId": garage_id,
            "roomName": room_name,
            "durationSeconds": duration_seconds,
            "transcript": transcript,
            "summary": summary,
            "confirmedBooking": confirmed_booking,
            "metrics": metrics,
        }

        # Only include optional string fields when non-empty (schema requires min 1 char if present)
        if customer_name:
            payload["customerName"] = customer_name
        if customer_phone:
            payload["customerPhone"] = customer_phone
        if registration_number:
            payload["registrationNumber"] = registration_number
        if booking_details:
            payload["bookingDetails"] = booking_details
        if call_type and call_type != "unknown":
            payload["callType"] = call_type
        
        # Add recording URL if available
        if RECORDING_BASE_URL:
            recording_url = f"{RECORDING_BASE_URL.rstrip('/')}/{room_name}.mp4"
            payload["recordingUrl"] = recording_url
            logger.info(f"[PORTAL] Including recording URL: {recording_url}")
        else:
            logger.warning("[PORTAL] RECORDING_BASE_URL not set; recordingUrl will be omitted")
        
        headers = {
            "Content-Type": "application/json",
            "X-Webhook-Secret": PORTAL_WEBHOOK_SECRET,
        }
        
        logger.info(f"[PORTAL] Posting call to {PORTAL_API_URL} | transcript={len(transcript)} entries | metrics_keys={list(metrics.keys())} | customerPhone={'YES' if customer_phone else 'OMITTED'}")
        
        async with aiohttp.ClientSession() as session:
            async with session.post(PORTAL_API_URL, json=payload, headers=headers) as response:
                if response.status == 201:
                    data = await response.json()
                    logger.info(f"[PORTAL] Call logged successfully: {data.get('callId', 'unknown')}")
                else:
                    text = await response.text()
                    logger.error(f"[PORTAL] Failed to log call: {response.status} - {text}")
    except Exception as e:
        logger.error(f"[PORTAL] Error logging call: {e}")

# ============================================================
# NATO / VRN NORMALIZATION
# ============================================================

_NATO_LETTER_MAP = {
    "alpha": "A", "alfa": "A",
    "bravo": "B",
    "charlie": "C",
    "delta": "D", "david": "D",
    "echo": "E", "edward": "E",
    "foxtrot": "F", "freddy": "F", "freddie": "F",
    "golf": "G", "george": "G",
    "hotel": "H", "harry": "H", "henry": "H",
    "india": "I",
    "juliet": "J", "juliett": "J",
    "kilo": "K", "king": "K",
    "lima": "L", "london": "L",
    "mike": "M", "mary": "M", "michael": "M",
    "november": "N", "nancy": "N", "nellie": "N",
    "oscar": "O", "oliver": "O",
    "papa": "P", "peter": "P",
    "quebec": "Q", "queen": "Q",
    "romeo": "R", "robert": "R", "roger": "R",
    "sierra": "S", "samuel": "S", "sugar": "S",
    "tango": "T", "tommy": "T", "thomas": "T",
    "uniform": "U", "uncle": "U",
    "victor": "V", "victory": "V", "victoria": "V",
    "whiskey": "W", "whisky": "W", "william": "W",
    "xray": "X", "x-ray": "X",
    "yankee": "Y", "yellow": "Y",
    "zulu": "Z", "zebra": "Z",
}

_DIGIT_WORD_MAP = {
    "zero": "0", "oh": "0", "owe": "0", "o": "0", "naught": "0", "nought": "0",
    "one": "1", "won": "1", "two": "2", "too": "2", "to": "2",
    "three": "3", "tree": "3", "four": "4", "for": "4",
    "five": "5", "fife": "5", "six": "6", "seven": "7",
    "eight": "8", "ate": "8", "nine": "9", "niner": "9",
}

_ALL_NATO_DIGIT_WORDS: dict[str, str] = {}
_ALL_NATO_DIGIT_WORDS.update(_NATO_LETTER_MAP)
_ALL_NATO_DIGIT_WORDS.update(_DIGIT_WORD_MAP)
_SORTED_NATO_DIGIT_WORDS: list[tuple[str, str]] = sorted(
    ((word, char) for word, char in _ALL_NATO_DIGIT_WORDS.items() if len(word) >= 2),
    key=lambda x: len(x[0]),
    reverse=True,
)


def vrm_to_phonetics(vrm: str) -> str:
    """Convert a UK VRM to NATO phonetic alphabet for clear readback.
    Example: 'AB12CDE' → 'Alpha Bravo One Two Charlie Delta Echo'
    """
    _PHONETIC_ALPHABET = {
        'A': 'Alpha', 'B': 'Bravo', 'C': 'Charlie', 'D': 'Delta',
        'E': 'Echo', 'F': 'Foxtrot', 'G': 'Golf', 'H': 'Hotel',
        'I': 'India', 'J': 'Juliet', 'K': 'Kilo', 'L': 'Lima',
        'M': 'Mike', 'N': 'November', 'O': 'Oscar', 'P': 'Papa',
        'Q': 'Quebec', 'R': 'Romeo', 'S': 'Sierra', 'T': 'Tango',
        'U': 'Uniform', 'V': 'Victor', 'W': 'Whiskey', 'X': 'X-ray',
        'Y': 'Yankee', 'Z': 'Zulu',
        '0': 'Zero', '1': 'One', '2': 'Two', '3': 'Three',
        '4': 'Four', '5': 'Five', '6': 'Six', '7': 'Seven',
        '8': 'Eight', '9': 'Nine'
    }
    
    phonetics = []
    for char in vrm.upper():
        if char in _PHONETIC_ALPHABET:
            phonetics.append(_PHONETIC_ALPHABET[char])
        else:
            phonetics.append(char)  # Fallback for unexpected chars
    
    return ' '.join(phonetics)


def _scan_nato_blob(text: str) -> list[str]:
    result: list[str] = []
    lower = text.lower()
    i = 0
    while i < len(lower):
        if not lower[i].isalnum():
            i += 1
            continue
        matched = False
        for word, char in _SORTED_NATO_DIGIT_WORDS:
            end = i + len(word)
            if end <= len(lower) and lower[i:end] == word:
                result.append(char)
                i = end
                matched = True
                break
        if not matched:
            ch = lower[i]
            # UK VRNs never use the letter O — always digit 0.
            # Also map single-char digit words that were filtered from the sorted list.
            if ch in ("o",):
                result.append("0")
            else:
                result.append(ch.upper())
            i += 1
    return result


_CAR_MAKE_MODEL_WORDS = {
    "land", "rover", "range", "landrover", "rangerover",
    "bmw", "audi", "ford", "vauxhall", "volkswagen", "vw", "toyota",
    "honda", "nissan", "mercedes", "benz", "mercedesbenz",
    "peugeot", "citroen", "renault", "fiat", "seat", "skoda",
    "hyundai", "kia", "mazda", "subaru", "suzuki", "mitsubishi",
    "volvo", "jaguar", "bentley", "porsche", "mini", "jeep",
    "lexus", "infiniti", "tesla", "mg", "dacia", "cupra",
    "evoque", "sportage", "corsa", "fiesta", "focus", "golf",
    "polo", "civic", "yaris", "qashqai", "tucson", "tiguan",
}


def normalize_vehicle_registration(reg: str) -> str:
    if not reg:
        return ""
    # Strip car make/model words that callers sometimes append (e.g. "P20ALA Land Rover")
    tokens_raw = re.split(r"[\s,;:/\\-_]+", reg.strip())
    # Also split camelCase blobs like "P20ALALandRover" → check for make words fused at the end
    if len(tokens_raw) == 1 and len(tokens_raw[0]) > 7:
        blob = tokens_raw[0]
        # Try to find where a make/model word starts in the blob
        lower_blob = blob.lower()
        for make_word in sorted(_CAR_MAKE_MODEL_WORDS, key=len, reverse=True):
            idx = lower_blob.find(make_word)
            if idx >= 4:  # At least 4 chars of actual VRN before the make word
                blob = blob[:idx]
                logger.info(f"[VRN] Stripped make/model from blob: '{reg}' → '{blob}'")
                break
        tokens_raw = [blob]
    tokens = re.split(r"[\s,;:/\\-_]+", " ".join(tokens_raw))
    converted: list[str] = []
    for token in tokens:
        if not token:
            continue
        cleaned = re.sub(r"[^A-Za-z0-9]", "", token)
        if not cleaned:
            continue
        lower = cleaned.lower()
        # Skip standalone make/model words (e.g. "Land" "Rover" as separate tokens)
        if lower in _CAR_MAKE_MODEL_WORDS and len(lower) > 2:
            continue
        if lower in _NATO_LETTER_MAP:
            converted.append(_NATO_LETTER_MAP[lower])
            continue
        if lower in _DIGIT_WORD_MAP:
            converted.append(_DIGIT_WORD_MAP[lower])
            continue
        if len(cleaned) == 1:
            converted.append(cleaned.upper())
            continue
        converted.extend(_scan_nato_blob(cleaned))
    if converted:
        return "".join(converted)
    return "".join(c.upper() for c in reg if c.isalnum())


# ============================================================
# UTILITIES
# ============================================================

def _current_uk_datetime() -> datetime:
    return datetime.now(ZoneInfo("Europe/London"))


def is_within_business_hours() -> bool:
    """Check if current UK time is within garage business hours."""
    if not GARAGE_HOURS:
        # No hours configured, assume always open
        return True
    
    now = _current_uk_datetime()
    day_name = now.strftime("%A").lower()  # monday, tuesday, etc.
    
    # Check if hours exist for this day
    day_hours = GARAGE_HOURS.get(day_name)
    if not day_hours:
        return False  # Day not configured, assume closed
    
    # Handle closed days
    if day_hours.get("closed") is True or day_hours.get("isClosed") is True:
        return False
    
    # Parse open and close times
    open_time_str = day_hours.get("open") or day_hours.get("openTime")
    close_time_str = day_hours.get("close") or day_hours.get("closeTime")
    
    if not open_time_str or not close_time_str:
        return False  # Missing times, assume closed
    
    try:
        # Parse times (format: "08:30" or "8:30")
        open_hour, open_min = map(int, open_time_str.split(":"))
        close_hour, close_min = map(int, close_time_str.split(":"))
        
        current_minutes = now.hour * 60 + now.minute
        open_minutes = open_hour * 60 + open_min
        close_minutes = close_hour * 60 + close_min
        
        return open_minutes <= current_minutes < close_minutes
    except (ValueError, AttributeError):
        logger.error(f"[HOURS_CHECK] Failed to parse times for {day_name}: open={open_time_str}, close={close_time_str}")
        return True  # On error, assume open to avoid blocking


def get_business_hours_text() -> str:
    """Generate human-readable business hours text from GARAGE_HOURS."""
    if not GARAGE_HOURS:
        return "Monday to Friday, half eight in the morning until six in the evening"
    
    # Group consecutive days with same hours
    days_order = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    hours_text_parts = []
    
    i = 0
    while i < len(days_order):
        day = days_order[i]
        day_hours = GARAGE_HOURS.get(day)
        
        if not day_hours or day_hours.get("closed") or day_hours.get("isClosed"):
            i += 1
            continue
        
        open_time = day_hours.get("open") or day_hours.get("openTime")
        close_time = day_hours.get("close") or day_hours.get("closeTime")
        
        if not open_time or not close_time:
            i += 1
            continue
        
        # Find consecutive days with same hours
        end_day_idx = i
        for j in range(i + 1, len(days_order)):
            next_day = days_order[j]
            next_hours = GARAGE_HOURS.get(next_day, {})
            if (next_hours.get("open") or next_hours.get("openTime")) == open_time and \
               (next_hours.get("close") or next_hours.get("closeTime")) == close_time and \
               not (next_hours.get("closed") or next_hours.get("isClosed")):
                end_day_idx = j
            else:
                break
        
        # Format day range
        if end_day_idx == i:
            day_range = days_order[i].capitalize()
        else:
            day_range = f"{days_order[i].capitalize()} to {days_order[end_day_idx].capitalize()}"
        
        # Format times (convert 24h to 12h with am/pm)
        def format_time(time_str):
            hour, minute = map(int, time_str.split(":"))
            if hour == 0:
                return f"12:{minute:02d}am"
            elif hour < 12:
                return f"{hour}:{minute:02d}am" if minute > 0 else f"{hour}am"
            elif hour == 12:
                return f"12:{minute:02d}pm" if minute > 0 else "12pm"
            else:
                return f"{hour-12}:{minute:02d}pm" if minute > 0 else f"{hour-12}pm"
        
        open_formatted = format_time(open_time)
        close_formatted = format_time(close_time)
        
        hours_text_parts.append(f"{day_range} {open_formatted} to {close_formatted}")
        i = end_day_idx + 1
    
    if not hours_text_parts:
        return "Monday to Friday, half eight in the morning until six in the evening"
    
    return ", ".join(hours_text_parts)


def get_dynamic_greeting(branch_name: str) -> str:
    if AGENT_GREETING_LINE:
        period = "morning"
        hour = _current_uk_datetime().hour
        if 12 <= hour < 17:
            period = "afternoon"
        elif 17 <= hour < 24:
            period = "evening"
        return re.sub(r"timeofday", f"good {period}", AGENT_GREETING_LINE, flags=re.IGNORECASE)
    # Use "garridge" phonetic spelling for British pronunciation
    return f"Hello, you're through to the garridge. How can I help?"


# ============================================================
# ERROR MONITORING
# ============================================================

class ErrorMonitor:
    @staticmethod
    async def send_discord_alert(
        error_msg: str, agent_name: str, room_name: str = "", extra: Optional[dict] = None,
    ) -> None:
        if not DISCORD_WEBHOOK_URL:
            return
        embed = {
            "title": f"ReceptionMate Error - {agent_name}",
            "description": error_msg[:2000],
            "color": 0xFF0000,
            "fields": [
                {"name": "Agent", "value": agent_name, "inline": True},
                {"name": "Room", "value": room_name or "N/A", "inline": True},
                {"name": "Time", "value": _current_uk_datetime().strftime("%Y-%m-%d %H:%M:%S"), "inline": True},
            ],
            "footer": {"text": "ReceptionMate Supervisor System"},
        }
        if extra:
            for k, v in extra.items():
                embed["fields"].append({"name": k, "value": str(v)[:1024], "inline": True})
        try:
            async with aiohttp.ClientSession() as http:
                async with http.post(
                    DISCORD_WEBHOOK_URL, json={"embeds": [embed]},
                    timeout=aiohttp.ClientTimeout(total=5),
                ) as resp:
                    if resp.status not in (200, 204):
                        logger.warning(f"[ERROR_MONITOR] Discord webhook returned {resp.status}")
        except Exception as e:
            logger.warning(f"[ERROR_MONITOR] Failed to send Discord alert: {e}")

    @staticmethod
    async def log_to_excel(
        error_msg: str, agent_name: str, room_name: str = "", error_type: str = "",
    ) -> None:
        try:
            await asyncio.to_thread(
                ErrorMonitor._write_excel_row, error_msg, agent_name, room_name, error_type,
            )
        except Exception as e:
            logger.warning(f"[ERROR_MONITOR] Failed to log to Excel: {e}")

    @staticmethod
    def _write_excel_row(error_msg: str, agent_name: str, room_name: str, error_type: str) -> None:
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            logger.warning("[ERROR_MONITOR] openpyxl not installed - skipping Excel logging")
            return
        path = ERROR_LOG_EXCEL_PATH
        if path.exists():
            wb = load_workbook(str(path))
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Error Log"
            ws.append(["Timestamp", "Agent", "Room", "Error Type", "Error Message"])
        ws.append([
            _current_uk_datetime().strftime("%Y-%m-%d %H:%M:%S"),
            agent_name, room_name or "N/A", error_type or "general", error_msg[:5000],
        ])
        wb.save(str(path))

    @staticmethod
    async def report_error(
        error_msg: str, agent_name: str, room_name: str = "",
        error_type: str = "", extra: Optional[dict] = None,
    ) -> None:
        await asyncio.gather(
            ErrorMonitor.send_discord_alert(error_msg, agent_name, room_name, extra),
            ErrorMonitor.log_to_excel(error_msg, agent_name, room_name, error_type),
        )


# ============================================================
# STATE MACHINE
# ============================================================

class Step(Enum):
    GREETING = "greeting"
    NEED_VRN = "need_vrn"
    CONFIRMING_VEHICLE = "confirming_vehicle"
    NEED_SERVICE = "need_service"
    NEED_TIMESLOT = "need_timeslot"
    NEED_CONTACT = "need_contact"
    CONFIRMED = "confirmed"
    DONE = "done"
    MESSAGE_ONLY = "message_only"


@dataclass
class CallState:
    step: Step = Step.GREETING
    intent: str = ""  # "new_booking", "quote", "message"
    service_hint: str = ""

    # Caller
    customer_name_first: str = ""
    customer_name_last: str = ""

    # Vehicle
    vrn: str = ""
    vrn_confirmed: bool = False
    session_id: str = ""
    vehicle_make: str = ""
    vehicle_model: str = ""

    # Service
    services_available: list[dict] = field(default_factory=list)
    service_selected_id: str = ""
    service_selected_name: str = ""
    service_price: str = ""
    
    # Tyre-specific info
    tyre_size: str = ""
    tyre_quality: str = ""  # "budget", "mid-range", or "premium"
    tyre_position: str = ""  # e.g., "front left", "both fronts", "all four"
    
    # Diagnostic notes (collected during questionnaire)
    diagnostic_notes: list[str] = field(default_factory=list)
    
    # Description of work for "Other" / vague service bookings (passed to GarageHive notes)
    other_service_description: str = ""

    # Timeslot
    timeslots_available: list[dict] = field(default_factory=list)
    booking_date: str = ""
    booking_time: str = ""

    # Contact
    contact_phone: str = ""
    contact_email: str = ""
    house_name_or_number: str = ""
    postcode: str = ""
    street: str = ""
    city: str = ""
    full_address: str = ""
    notes: str = ""

    # Message
    message: str = ""
    preferred_callback_time: str = ""

    # Message summary (from specialist LLM)
    message_summary: Optional[dict] = None

    # Tracking
    vrn_attempts: int = 0
    vrn_readback_rejections: int = 0  # tracks how many times caller said "No" to readback
    vrn_partial: str = ""  # accumulates partial VRN segments while caller spells it out
    vrn_pending: str = ""  # normalized VRN awaiting caller confirmation before API lookup
    booking_submit_pending: bool = False  # True when submit_booking failed and needs retry
    recent_transcripts: list[str] = field(default_factory=list)
    conversation_items: list[dict] = field(default_factory=list)  # full agent+customer turns for GPT summary


# ============================================================
# API CLIENT (shared aiohttp session)
# ============================================================

class GHClient:
    def __init__(self):
        self._session: Optional[aiohttp.ClientSession] = None
        self._instance = GH_CUSTOMER_ID
        self._location_id = int(GH_LOCATION_ID) if GH_LOCATION_ID else 23
        self._base = f"https://onlinebooking.garagehive.co.uk/api/external-booking/{self._instance}"

    async def _get_session(self) -> aiohttp.ClientSession:
        if self._session is None or self._session.closed:
            headers = {"Content-Type": "application/json"}
            if GH_API_KEY:
                headers["Authorization"] = f"Bearer {GH_API_KEY}"
            self._session = aiohttp.ClientSession(
                headers=headers,
                timeout=aiohttp.ClientTimeout(total=20),
            )
        return self._session

    async def init_and_set_vehicle(self, reg: str) -> dict:
        """Init booking + set vehicle info in one call."""
        s = await self._get_session()

        # Step 1: Init
        async with s.post(f"{self._base}/init") as resp:
            if resp.status >= 400:
                raw = await resp.text()
                logger.error(f"[GH] Init failed {resp.status}: {raw}")
                return {"error": f"Init failed (HTTP {resp.status})"}
            try:
                data = json.loads(await resp.text())
            except json.JSONDecodeError:
                return {"error": "Init response not JSON"}
            booking = data.get("booking", {})
            session_id = booking.get("session_id") or data.get("sessionId")
            if not session_id:
                return {"error": "No session_id in init response"}

        # Step 2: Set vehicle
        async with s.post(
            f"{self._base}/{session_id}/set-vehicle-info",
            json={"registration_no": reg, "reg_no_country": "GB", "location_id": self._location_id},
        ) as resp:
            vehicle_data = await resp.json()
            vehicle_data["session_id"] = session_id
            logger.info(f"[GH] init_and_set_vehicle success for {reg}")
            return vehicle_data

    async def list_services(self, session_id: str) -> list[dict]:
        s = await self._get_session()
        async with s.get(f"{self._base}/{session_id}/list-services") as resp:
            data = await resp.json()
            return data.get("services") or []

    async def set_service(self, session_id: str, service_price_ids: str) -> dict:
        s = await self._get_session()
        raw_ids = [p.strip() for p in str(service_price_ids).split(",") if p.strip()]
        ids: list[int | str] = [int(x) if x.isdigit() else x for x in raw_ids]
        async with s.post(
            f"{self._base}/{session_id}/set-services",
            json={"servicePriceIDs": ids},
        ) as resp:
            return await resp.json()

    async def list_timeslots(self, session_id: str) -> list[dict]:
        s = await self._get_session()
        async with s.get(f"{self._base}/{session_id}/list-timeslots") as resp:
            data = await resp.json()
            timeslots = data.get("timeslots") or {}
            result = []
            for dt, times in timeslots.items():
                for t in times:
                    result.append({"date": dt, "time": t})
            return result

    async def set_timeslot(self, session_id: str, date: str, time: str) -> dict:
        s = await self._get_session()
        async with s.post(
            f"{self._base}/{session_id}/set-timeslot",
            json={"bookingDate": date, "bookingTime": time},
        ) as resp:
            return await resp.json()

    async def set_contact_info(self, session_id: str, **kwargs) -> dict:
        s = await self._get_session()
        async with s.post(
            f"{self._base}/{session_id}/set-contact-info",
            json=kwargs,
        ) as resp:
            http_status = resp.status
            try:
                data = await resp.json()
                if 200 <= http_status < 300 or (isinstance(data, dict) and data.get("status") == "success"):
                    return {"status": "success", "booking": data.get("booking", {})}
                return {
                    "status": "error",
                    "message": data.get("message", "Failed to confirm booking"),
                    "errors": data.get("errors", []),
                }
            except Exception:
                if 200 <= http_status < 300:
                    return {"status": "success", "booking": {}}
                return {"status": "error", "message": f"HTTP {http_status}"}

    async def validate_address(self, postcode: str) -> dict:
        clean = postcode.replace(" ", "").upper()
        try:
            timeout = aiohttp.ClientTimeout(total=5)
            async with aiohttp.ClientSession(timeout=timeout) as s:
                async with s.get(f"https://api.postcodes.io/postcodes/{clean}") as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        if data.get("status") == 200 and data.get("result"):
                            r = data["result"]
                            return {
                                "street": r.get("parish") or r.get("admin_ward") or "",
                                "city": r.get("admin_district") or r.get("postcode_area") or "",
                            }
        except Exception as e:
            logger.warning(f"[ADDRESS] Postcode lookup failed: {e}")
        return {"street": "", "city": ""}

    async def close(self):
        if self._session and not self._session.closed:
            await self._session.close()


# ============================================================
# PHONE NUMBER SANITISATION
# ============================================================

# STT mishearings of "plus" — Deepgram transcribes "plus 44" as "flush 44", "blush 44", etc.
_PLUS_MISHEARINGS = ("flush", "blush", "plush", "plus", "flash")


def _sanitise_phone(raw: str) -> str:
    """Clean a phone number that may contain STT artefacts.
    Handles both word-form ('flush four four eight hundred...') and
    pre-concatenated ('flush448002062757') from the LLM.
    """
    if not raw:
        return raw
    cleaned = raw.strip().lower()
    # Replace "flush"/"blush"/etc. at the start with "+"
    for word in _PLUS_MISHEARINGS:
        if cleaned.startswith(word):
            cleaned = "+" + cleaned[len(word):]
            break
    # Strip everything except digits and leading +
    digits = re.sub(r"[^0-9+]", "", cleaned)
    # Ensure at most one + at the start
    if "+" in digits:
        digits = "+" + digits.replace("+", "")
    return digits or raw  # return original if sanitisation emptied it


# ============================================================
# EMAIL SANITISATION
# ============================================================

# Spoken number words → digits (LLM may pass "one two three" instead of "123")
_SPOKEN_DIGITS = {
    "zero": "0", "one": "1", "two": "2", "three": "3", "four": "4",
    "five": "5", "six": "6", "seven": "7", "eight": "8", "nine": "9",
}

# Common STT/LLM artefacts in email addresses
_EMAIL_WORD_MAP = {
    "underscore": "_", "under score": "_", "dash": "-", "hyphen": "-",
    "dot": ".", "full stop": ".", "period": ".", "at sign": "@", " at ": "@",
}


def _sanitise_email(raw: str) -> str:
    """Clean an email that may contain spoken digit words or STT artefacts.
    'gabriel_one_two_three@gmail.com' → 'gabriel_123@gmail.com'
    """
    if not raw or "@" not in raw:
        return raw
    cleaned = raw.strip().lower().replace(" ", "")

    # Replace spoken digit words with actual digits
    for word, digit in _SPOKEN_DIGITS.items():
        cleaned = cleaned.replace(word, digit)

    # Collapse _digit_digit_ sequences → contiguous digits
    # "gabriel_1_2_3@" → "gabriel_123@"
    while re.search(r"(\d)[_](\d)", cleaned):
        cleaned = re.sub(r"(\d)[_](\d)", r"\1\2", cleaned)

    cleaned = cleaned.replace(" ", "")
    return cleaned


# ============================================================
# SERVICE MATCHING
# ============================================================

def _normalize_service_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


# ── Service Advisor Specialist (replaces keyword rules) ──
# Uses GPT-4o-mini via LiveKit Cloud inference with a focused
# automotive prompt to match vague caller descriptions to services.

_SERVICE_ADVISOR_PROMPT = """\
You are an automotive service advisor at a UK garage.
Given the customer's description and the available services, pick the single most suitable service.

Rules:
- "hasn't been serviced in ages/long time/overdue" → Full Service
- Noises, rattles, warning lights, unknown issues → Diagnostic Check
- Specific systems (brakes, oil, tyres, air con, cam belt) → match to the relevant service
- MOT/test → MOT
- BRAKE-RELATED: Be very careful:
  * "brake fluid" / "brake fluid change" → Brake Fluid Change
  * "brake pads" / "pads" / "brake pad replacement" → Brake Pad Replacement (or Front/Rear if specified)
  * "brake change" / "brakes" (vague) → return null (too ambiguous - could be fluid, pads, discs, or full service)
  * "brake discs" / "rotors" → Brake Disc Replacement
- If genuinely unclear or ambiguous, return null

Reply with JSON ONLY — no extra text:
{"service_name": "exact name from the list", "reason": "one short sentence for the receptionist to say"}\
"""


def _format_price(svc: dict) -> str:
    """Return a spoken price string for a GarageHive service dict.

    Rules:
      hide_service_prices=true  → "" (caller never hears a price)
      price=0 (or missing)      → "" (available on request, don't quote £0)
      from_price=true           → "from £X"
      estimate=true             → "from around £X"
      otherwise                 → "£X"
    """
    if svc.get("hide_service_prices"):
        return ""
    price = svc.get("price", 0) or 0
    if not price:
        return ""
    if svc.get("estimate"):
        return f"from around £{price}"
    if svc.get("from_price"):
        return f"from £{price}"
    return f"£{price}"


async def specialist_service_match(
    caller_text: str, services: list[dict]
) -> Optional[tuple[dict, str]]:
    """Ask the Service Advisor specialist LLM to match a vague description to a service.
    Returns (matched_service_dict, reason_string) or None.
    Falls back gracefully on error/timeout."""
    if not services or not caller_text:
        return None

    svc_list = "\n".join(
        f"- {s.get('name', '?')}{(' (' + _format_price(s) + ')') if _format_price(s) else ''}" for s in services
    )
    user_msg = f"Customer said: \"{caller_text}\"\n\nAvailable services:\n{svc_list}"

    try:
        client = _get_specialist_llm()
        if client is None:
            logger.debug("[SERVICE_ADVISOR] No LiveKit credentials — skipping specialist, using fuzzy match only")
            return None
        resp = await asyncio.wait_for(
            client.chat.completions.create(
                model="openai/gpt-4o-mini",
                temperature=0.1,
                max_tokens=150,
                messages=[
                    {"role": "system", "content": _SERVICE_ADVISOR_PROMPT},
                    {"role": "user", "content": user_msg},
                ],
            ),
            timeout=5.0,
        )
        raw = resp.choices[0].message.content.strip()
        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
        data = json.loads(raw)

        svc_name = data.get("service_name")
        reason = data.get("reason", "")
        if not svc_name or svc_name == "null":
            return None

        # Match the returned name to the actual service dict
        for svc in services:
            if svc.get("name", "").lower() == svc_name.lower():
                logger.info(f"[SERVICE_ADVISOR] Matched: '{svc_name}' — {reason}")
                return (svc, reason)

        # Fuzzy fallback: the specialist might return a slightly different name
        svc_lower = svc_name.lower()
        for svc in services:
            if svc_lower in svc.get("name", "").lower() or svc.get("name", "").lower() in svc_lower:
                logger.info(f"[SERVICE_ADVISOR] Fuzzy matched: '{svc_name}' → '{svc.get('name')}' — {reason}")
                return (svc, reason)

        logger.warning(f"[SERVICE_ADVISOR] LLM suggested '{svc_name}' but no match in services list")
        return None

    except asyncio.TimeoutError:
        logger.warning("[SERVICE_ADVISOR] Timed out (5s) — falling back to fuzzy match")
        return None
    except Exception as e:
        logger.warning(f"[SERVICE_ADVISOR] Error: {e} — falling back to fuzzy match")
        return None


# ── Diagnostic Questionnaire Specialist ──
# Detects when caller describes a symptom/fault and generates appropriate diagnostic questions.
# Follows the structured fault diagnostic flow.

_DIAGNOSTIC_QUESTIONNAIRE_PROMPT = """\
You are a diagnostic specialist at a UK garage following a structured fault diagnosis process.
The customer has described a symptom or problem with their vehicle.

DIAGNOSTIC FLOW:
1. STEP 1 (if description is vague): Ask the broad open question: "Can you tell me what it's doing?"
2. STEP 2: Clarify the symptom type based on what they've said:
   
   🚨 NOISE: Ask:
   - "When does it happen — when driving, braking, starting?"
   - "Is it constant or intermittent?"
   - "Does it change with speed?"
   
   ⚠️ WARNING LIGHT: Ask:
   - "Which light is on — engine, ABS, battery?"
   - "Is the car driving normally?"
   - "Has it gone into limp mode?"
   
   🛑 PERFORMANCE ISSUE: Ask:
   - "Is it struggling to start?"
   - "Loss of power?"
   - "Cutting out at all?"
   
   🌡 OVERHEATING/SMELL/SMOKE: Ask:
   - "Any smoke at all?"
   - "Any burning smell?"
   - "Has the temperature gauge gone high?"

3. STEP 3 (timing - ALWAYS ask): 
   - "When did this first start?"
   - "Has it got worse?"

Rules:
- Keep questions SHORT and conversational
- Use UK English: "when driving, braking" not "during operation"
- Generate questions in order: symptom clarification first, then timing
- Max 4 questions total
- If description is vague, start with "Can you tell me what it's doing?"

Reply with JSON ONLY:
{"questions": ["question 1", "question 2", "question 3"], "symptom_type": "noise|warning_light|performance|overheating|vague"}
or
{"questions": [], "symptom_type": "detailed_already"}\
"""


async def specialist_diagnostic_questions(
    symptom_description: str,
) -> Optional[dict]:
    """Ask the Diagnostic specialist to generate follow-up questions for a symptom.
    Returns dict with questions list and symptom_type, or None.
    Falls back gracefully on error/timeout."""
    if not symptom_description:
        return None

    # Check if this is actually a symptom/fault (contains keywords)
    symptom_keywords = [
        "noise", "sound", "knock", "rattle", "squeal", "grind", "click", "clunk",
        "vibrat", "shak", "judder", "pull",
        "warning", "light", "dashboard", "check engine", "abs", "battery",
        "smell", "smoke", "leak", "overheat", "hot", "burning",
        "problem", "issue", "fault", "wrong", "broken",
        "not working", "won't", "doesn't", "can't", "struggling",
        "rough", "harsh", "stuttering", "hesitat", "cutting out",
        "loss of power", "lost power", "losing power", "no power", "lack of power",
        "limp mode", "stall", "misfire", "bog", "sluggish",
    ]
    
    symptom_lower = symptom_description.lower()
    if not any(keyword in symptom_lower for keyword in symptom_keywords):
        logger.debug(f"[DIAGNOSTIC_Q] No symptom keywords detected in: '{symptom_description}'")
        return None

    user_msg = f"Customer description: \"{symptom_description}\""

    try:
        client = _get_specialist_llm()
        if client is None:
            logger.debug("[DIAGNOSTIC_Q] No LiveKit credentials — skipping diagnostic questions")
            return None
        
        resp = await asyncio.wait_for(
            client.chat.completions.create(
                model="openai/gpt-4o-mini",
                temperature=0.2,
                max_tokens=250,
                messages=[
                    {"role": "system", "content": _DIAGNOSTIC_QUESTIONNAIRE_PROMPT},
                    {"role": "user", "content": user_msg},
                ],
            ),
            timeout=5.0,
        )
        
        raw = resp.choices[0].message.content.strip()
        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
        data = json.loads(raw)

        questions = data.get("questions", [])
        symptom_type = data.get("symptom_type", "unknown")
        
        if not questions:
            logger.info(f"[DIAGNOSTIC_Q] No questions needed — symptom_type: {symptom_type}")
            return None

        logger.info(f"[DIAGNOSTIC_Q] Generated {len(questions)} questions for {symptom_type}")
        return {"questions": questions, "symptom_type": symptom_type}

    except asyncio.TimeoutError:
        logger.warning("[DIAGNOSTIC_Q] Timed out (5s) — skipping diagnostic questions")
        return None
    except Exception as e:
        logger.warning(f"[DIAGNOSTIC_Q] Error: {e} — skipping diagnostic questions")
        return None


# ── Timeslot Matcher Specialist ──
# Parses natural language timeslot preferences against available slots.
# Deterministic at temp 0 for consistent date parsing.

_TIMESLOT_MATCHER_PROMPT = """\
You are a scheduling assistant at a UK garage.
Given today's date, the caller's preference, and the available timeslots, pick the single best matching slot.

Rules:
- "morning" = before 12:00, "afternoon" = 12:00-17:00
- "tomorrow" = the day after today
- "next week" = Monday of the following week onward
- "ASAP" / "first available" / "earliest" = the earliest slot in the list
- "the first one" / "the morning one" = match to the first/morning slot from those offered
- If the caller picks a specific slot, match it exactly
- CRITICAL: If the caller asks for a SPECIFIC DAY (e.g. "Friday", "Monday") and NO slot exists on that day, return null. NEVER substitute a different day — the receptionist will offer alternatives.
- If genuinely unclear or no available slot matches, return null

Reply with JSON ONLY:
{"date": "YYYY-MM-DD", "time": "HH:MM"}
or
{"date": null, "time": null}\
"""


async def specialist_timeslot_match(
    caller_preference: str, available_slots: list[dict], today: str
) -> Optional[dict]:
    """Ask the Timeslot Matcher specialist to parse caller's preference.
    Returns {"date": "...", "time": "..."} or None."""
    if not caller_preference or not available_slots:
        return None

    slot_list = "\n".join(
        f"- {s.get('date', '?')} at {s.get('time', '?')}" for s in available_slots[:30]
    )
    user_msg = f"Today is {today}.\nCaller said: \"{caller_preference}\"\n\nAvailable timeslots:\n{slot_list}"

    try:
        client = _get_specialist_llm()
        if client is None:
            logger.debug("[TIMESLOT_MATCHER] No LiveKit credentials — skipping specialist")
            return None
        resp = await asyncio.wait_for(
            client.chat.completions.create(
                model="openai/gpt-4o-mini",
                temperature=0,
                max_tokens=80,
                messages=[
                    {"role": "system", "content": _TIMESLOT_MATCHER_PROMPT},
                    {"role": "user", "content": user_msg},
                ],
            ),
            timeout=5.0,
        )
        raw = resp.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
        data = json.loads(raw)

        date_val = data.get("date")
        time_val = data.get("time")
        if not date_val or date_val == "null" or not time_val or time_val == "null":
            return None

        # Verify the slot actually exists in the available list
        for slot in available_slots:
            if slot.get("date") == date_val and slot.get("time") == time_val:
                logger.info(f"[TIMESLOT_MATCHER] Matched: {date_val} at {time_val}")
                return {"date": date_val, "time": time_val}

        logger.warning(f"[TIMESLOT_MATCHER] LLM suggested {date_val} {time_val} but not in available slots")
        return None

    except asyncio.TimeoutError:
        logger.warning("[TIMESLOT_MATCHER] Timed out (5s)")
        return None
    except Exception as e:
        logger.warning(f"[TIMESLOT_MATCHER] Error: {e}")
        return None


# ── Message Summariser Specialist ──
# Structures raw caller messages into category + summary + action for the team.

_MESSAGE_SUMMARISER_PROMPT = """\
You are a receptionist summarising a caller's message for the garage team to action when calling back.
Be concise and professional.

Reply with JSON ONLY:
{"category": "vehicle_update|complaint|enquiry|callback|urgent|other", "summary": "2-3 sentence summary", "action": "what the team should do when calling back"}\
"""


async def specialist_summarise_message(
    raw_message: str, caller_name: str, vehicle: str = ""
) -> Optional[dict]:
    """Summarise a caller's message for the team.
    Returns {"category": "...", "summary": "...", "action": "..."} or None."""
    if not raw_message:
        return None

    user_msg = f"Caller: {caller_name}\nVehicle: {vehicle or 'not provided'}\nMessage: \"{raw_message}\""

    try:
        client = _get_specialist_llm()
        if client is None:
            logger.debug("[MESSAGE_SUMMARISER] No LiveKit credentials — skipping specialist")
            return None
        resp = await asyncio.wait_for(
            client.chat.completions.create(
                model="openai/gpt-4o-mini",
                temperature=0.2,
                max_tokens=200,
                messages=[
                    {"role": "system", "content": _MESSAGE_SUMMARISER_PROMPT},
                    {"role": "user", "content": user_msg},
                ],
            ),
            timeout=5.0,
        )
        raw = resp.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
        data = json.loads(raw)

        if data.get("summary"):
            logger.info(f"[MESSAGE_SUMMARISER] Category: {data.get('category')}, Summary: {data.get('summary')[:80]}")
            return data
        return None

    except asyncio.TimeoutError:
        logger.warning("[MESSAGE_SUMMARISER] Timed out (5s)")
        return None
    except Exception as e:
        logger.warning(f"[MESSAGE_SUMMARISER] Error: {e}")
        return None


def match_service(name_hint: str, services: list[dict]) -> Optional[dict]:
    """Fuzzy-match a caller's service description to the best available service."""
    if not services or not name_hint:
        return None
    target = _normalize_service_text(name_hint)
    if not target:
        return None

    best_match: Optional[dict] = None
    best_score = 0.0
    for service in services:
        svc_name = _normalize_service_text(service.get("name", ""))
        if not svc_name:
            continue
        if svc_name == target:
            return service
        score = SequenceMatcher(None, target, svc_name).ratio()
        if target in svc_name or svc_name in target:
            score += 0.25
        if score > best_score:
            best_score = score
            best_match = service
    return best_match if best_score >= 0.45 else None


# ============================================================
# SUPERVISOR AGENT
# ============================================================

class SupervisorAgent(Agent):
    """Single speaking agent that manages the entire call via worker tools."""

    def __init__(self, state: CallState, gh: GHClient, room_name: str = "", assist_mode: bool = False):
        self._state = state
        self._gh = gh
        self._room_name = room_name
        self._agent_session: Optional[AgentSession] = None
        self._assist_mode = assist_mode  # If True, don't speak first greeting

        # ── Worker Tools ─────────────────────────────────────

        @function_tool
        async def get_current_datetime(context: RunContext) -> dict:
            """Get the current UK date and time."""
            now = _current_uk_datetime()
            return {"date": now.strftime("%A %d %B %Y"), "time": now.strftime("%I:%M %p")}

        @function_tool
        async def save_caller_name(
            context: RunContext,
            first_name: str,
            last_name: str = "",
            intent: str = "",
            service_hint: str = "",
            vrn: str = "",
            requested_person: str = "",
        ) -> str:
            """Save the caller's name and determine intent.
            intent: 'booking' for appointments, 'quote' for price enquiries,
                    'vehicle_update' for callers checking on a vehicle already at the garage,
                    'message' for everything else,
                    'transfer' for asking to speak to someone specific or a human.
            service_hint: what work they mentioned (e.g. 'MOT', 'full service').
            vrn: vehicle registration if the caller already gave it.
            requested_person: name of person they asked for (e.g. 'John', 'manager', 'human')."""

            if self._state.step != Step.GREETING:
                return f"ERROR: Wrong step ({self._state.step.value}). Cannot save name now."

            first = (first_name or "").strip()
            last = (last_name or "").strip()
            requested = (requested_person or "").strip()

            # Detect stutter/correction: "My name is Gab Gabriel" → first="Gab", last="Gabriel"
            # If first is a prefix of last, the caller stuttered — use last as the actual first name.
            if first and last and first.lower() != last.lower():
                if last.lower().startswith(first.lower()) and len(first) < len(last):
                    logger.info(f"[SAVE_NAME] Detected stutter: '{first}' + '{last}' → using '{last}' as first name")
                    first = last
                    last = ""

            # Guard: reject single-character or implausibly short names
            if first and len(first) < 2:
                return (
                    f"REJECTED: '{first}' is too short to be a name. "
                    "ASK the caller: 'Can I take your name?' and WAIT for their response."
                )

            # Guard: verify the caller actually said this name (as a whole word)
            all_speech = " ".join(self._state.recent_transcripts).lower()
            # Check as whole word to avoid matching "v" inside "service"
            speech_words = set(re.findall(r"[a-z]+", all_speech))
            if first and first.lower() not in speech_words:
                return (
                    f"REJECTED: The caller has NOT said the name '{first}'. "
                    "You are hallucinating a name. ASK the caller: 'Can I take your name?' "
                    "and WAIT for their response. Only call save_caller_name with the name they actually say."
                )

            self._state.customer_name_first = first
            self._state.customer_name_last = last

            resolved = (intent or "").strip().lower()
            logger.info(f"[SAVE_NAME] {first} {last}, intent={resolved}, hint={service_hint}, vrn={vrn}, requested_person={requested}")

            # Transfer request - asking for a specific person or human
            if resolved in ("transfer", "speak_to", "ask_for") or requested:
                self._state.intent = "message"
                self._state.step = Step.MESSAGE_ONLY
                person_mention = f" for {requested}" if requested else ""
                
                # Check if we're outside business hours
                if not is_within_business_hours():
                    return (
                        f"Name saved: {first} {last}. Intent: transfer request{person_mention}.\n"
                        f"Address the caller as '{first}' (FIRST name only).\n"
                        f"OUTSIDE BUSINESS HOURS. Say naturally: 'The team aren't available outside of our opening hours, "
                        f"but I can take a message and they'll give you a ring back when we're open. What would you like me to pass on?'\n"
                        f"Then collect message details with take_message."
                    )
                
                return (
                    f"Name saved: {first} {last}. Intent: transfer request{person_mention}.\n"
                    f"Address the caller as '{first}' (FIRST name only).\n"
                    f"Say naturally: 'Unfortunately the team aren't available at the moment — they're likely helping other customers. "
                    f"However, I can help you with bookings, or I can take a message and get someone to give you a ring back. Which would you prefer?'\n"
                    f"If they want a booking → switch to booking flow (ask what work they need).\n"
                    f"If they want a message → ask 'What would you like the team to know?' then collect phone number and take_message."
                )

            # Message path
            if resolved in ("message", "enquiry", "reschedule", "cancel", "complaint", "question"):
                self._state.intent = "message"
                self._state.step = Step.MESSAGE_ONLY
                return (
                    f"Name saved: {first} {last}. Intent: message.\n"
                    f"Address the caller as '{first}' (FIRST name only — never use their surname to greet them).\n"
                    "Now take their message. Ask: 'What would you like the team to know?'\n"
                    "Then collect their phone number and preferred callback time."
                )

            # Vehicle update path — caller wants status on a vehicle already at the garage
            if resolved in ("vehicle_update", "update", "status", "progress"):
                self._state.intent = "vehicle_update"
                
                # Check if we're outside business hours
                if not is_within_business_hours():
                    self._state.step = Step.MESSAGE_ONLY
                    return (
                        f"Name saved: {first} {last}. Intent: vehicle update.\n"
                        f"Address the caller as '{first}' (FIRST name only).\n"
                        f"OUTSIDE BUSINESS HOURS. Say naturally: 'The team aren't available outside of our opening hours to check on your vehicle, "
                        f"but I can take a message and they'll give you a ring back when we're open. What would you like me to pass on?'\n"
                        f"Then collect message details with take_message."
                    )
                
                if vrn:
                    self._state.step = Step.NEED_VRN
                    return (
                        f"Name saved: {first} {last}. Intent: vehicle update.\n"
                        f"Address the caller as '{first}' (FIRST name only).\n"
                        f"Caller provided registration: '{vrn}'.\n"
                        f"NOW call lookup_vehicle(reg='{vrn}') to parse it. GENERATE ZERO SPEECH."
                    )
                self._state.step = Step.NEED_VRN
                return (
                    f"Name saved: {first} {last}. Intent: vehicle update.\n"
                    f"Address the caller as '{first}' (FIRST name only).\n"
                    "Say: 'No problem. Could I grab your registration so I can pull up your details?'\n"
                    "Then STOP. Wait for the caller to respond."
                )

            # Booking / quote path
            self._state.intent = "quote" if resolved == "quote" else "new_booking"
            if service_hint:
                self._state.service_hint = service_hint.strip()

            # If caller provided both VRN and indicated they want a booking, proceed
            if vrn and (service_hint or resolved == "booking"):
                self._state.step = Step.NEED_VRN
                return (
                    f"Name saved: {first} {last}. Address caller as '{first}' (FIRST name only).\n"
                    f"Caller provided registration: '{vrn}'.\n"
                    f"NOW call lookup_vehicle(reg='{vrn}') to parse it. GENERATE ZERO SPEECH."
                )

            # If no intent was clearly stated, establish the reason for the call first
            if not resolved and not service_hint:
                self._state.step = Step.GREETING  # Stay in greeting until we know why they're calling
                return (
                    f"Name saved: {first} {last}. Address caller as '{first}' (FIRST name only).\n"
                    "ESTABLISH REASON: The caller hasn't said why they're calling.\n"
                    "Ask naturally: 'How can I help you today?' or 'What can I do for you?'\n"
                    "Wait for them to explain (booking, quote, question, etc.), then call save_caller_name again with the intent/service_hint."
                )

            # If they indicated booking/quote but no VRN yet
            if vrn:
                self._state.step = Step.NEED_VRN
                return (
                    f"Name saved: {first} {last}. Address caller as '{first}' (FIRST name only).\n"
                    f"Caller provided registration: '{vrn}'.\n"
                    f"NOW call lookup_vehicle(reg='{vrn}') to parse it. GENERATE ZERO SPEECH."
                )

            self._state.step = Step.NEED_VRN
            return (
                f"Name saved: {first} {last}. Address caller as '{first}' (FIRST name only).\n"
                "Say EXACTLY ONE short sentence asking for their registration, e.g. 'Could I grab your registration?'\n"
                "Then STOP. Generate NOTHING else. Wait for the caller to respond."
            )

        @function_tool
        async def lookup_vehicle(context: RunContext, reg: str, confirmed: bool = False) -> str:
            """Look up a vehicle by UK registration number.
            
            TWO-STEP process:
            1. Pass caller's words (confirmed=False) → tool normalizes and gives you phonetic readback
            2. After caller confirms → call with confirmed=True to do API lookup
            
            The tool handles NATO phonetic conversion automatically."""

            if self._state.step not in (Step.NEED_VRN, Step.CONFIRMING_VEHICLE, Step.MESSAGE_ONLY):
                if self._state.step == Step.GREETING:
                    return (
                        "BLOCKED: You must call save_caller_name FIRST before any other tool. "
                        "Get the caller's name and intent, then call save_caller_name."
                    )
                return f"ERROR: Wrong step ({self._state.step.value}). Vehicle lookup not needed now."

            # Intent recovery: caller was on message path but actually wants to book
            if self._state.step == Step.MESSAGE_ONLY:
                self._state.intent = "new_booking"
                self._state.step = Step.NEED_VRN
                logger.info("[LOOKUP] Intent recovery: MESSAGE_ONLY → NEED_VRN")

            # ── STEP 2: Caller confirmed — do the actual API lookup ──
            if confirmed:
                normalized = self._state.vrn_pending or normalize_vehicle_registration(reg)
                self._state.vrn_pending = ""
                self._state.vrn = normalized
                self._state.vrn_attempts += 1
                
                logger.info(f"[LOOKUP] Caller confirmed '{normalized}' — API lookup (attempt {self._state.vrn_attempts})")

                # Try automatic B↔V↔P correction (common misheard letters)
                _BVP_SWAPS = {"B": ["V", "P"], "V": ["B", "P"], "P": ["B", "V"]}
                regs_to_try = [normalized]
                first_char = normalized[0] if normalized else ""
                if first_char in _BVP_SWAPS:
                    for alt in _BVP_SWAPS[first_char]:
                        regs_to_try.append(alt + normalized[1:])

                result = None
                winning_reg = normalized
                for try_reg in regs_to_try:
                    try:
                        attempt_result = await self._gh.init_and_set_vehicle(try_reg)
                    except Exception as e:
                        logger.error(f"[LOOKUP] API error for '{try_reg}': {e}")
                        continue

                    if "error" in attempt_result:
                        logger.info(f"[LOOKUP] Not found: '{try_reg}'")
                        continue

                    booking_data = attempt_result.get("booking", {})
                    vehicle_data = booking_data.get("vehicle", {})
                    if vehicle_data.get("make_name") or vehicle_data.get("model_name"):
                        result = attempt_result
                        winning_reg = try_reg
                        if try_reg != normalized:
                            logger.info(f"[LOOKUP] B/V/P auto-fix: '{normalized}' → '{try_reg}'")
                        break

                # Vehicle not found
                if result is None:
                    asyncio.create_task(ErrorMonitor.report_error(
                        error_msg=f"lookup_vehicle failed for: {normalized}",
                        agent_name="SUPERVISOR",
                        room_name=self._room_name,
                        error_type="api_error",
                        extra={"reg": normalized, "attempt": self._state.vrn_attempts},
                    ))
                    
                    if self._state.vrn_attempts >= 3:
                        self._state.step = Step.MESSAGE_ONLY
                        return (
                            "Vehicle not found after 3 attempts.\n"
                            "Say: 'I'm having trouble finding that one. Let me take your details "
                            "and get the team to ring you back.'\n"
                            "Collect message, phone number, then call take_message."
                        )
                    
                    return (
                        f"Vehicle not found for '{normalized}'.\n"
                        "Say: 'I'm not finding that one. Could you read it out again, letter by letter?'\n"
                        "When they provide it again, call lookup_vehicle with confirmed=false."
                    )

                # Extract vehicle info
                normalized = winning_reg
                self._state.vrn = normalized
                booking = result.get("booking", {})
                vehicle = booking.get("vehicle", {})
                make = vehicle.get("make_name", "")
                model = vehicle.get("model_name", "")
                session_id = result.get("session_id", "")

                if not make and not model:
                    return "Vehicle data empty. Ask them to spell the registration again."

                self._state.session_id = session_id
                self._state.vehicle_make = make
                self._state.vehicle_model = model
                self._state.step = Step.CONFIRMING_VEHICLE

                first = self._state.customer_name_first
                last = self._state.customer_name_last
                logger.info(f"[LOOKUP] Found: {make} {model}, session={session_id}")

                return (
                    f"Vehicle found: {make.title()} {model.title()} (registration {normalized}).\n"
                    f"Say: 'I've got a {make.title()} {model.title()} on that registration. Is that right?'\n"
                    "Wait for YES/NO.\n"
                    f"If YES → confirm name: 'And I've got your name as {first} {last} — is that right?'\n"
                    "If both confirmed → call confirm_vehicle(confirmed=true).\n"
                    "If name needs correction → call confirm_vehicle(confirmed=true, corrected_first_name='...', corrected_last_name='...').\n"
                    "If VEHICLE is wrong → call confirm_vehicle(confirmed=false)."
                )

            # ── STEP 1: Normalize VRN and generate phonetic readback ──
            normalized = normalize_vehicle_registration(reg)
            
            logger.info(f"[LOOKUP] Normalizing '{reg}' → '{normalized}'")
            
            # Validation: minimum length
            if len(normalized) < 4:
                return (
                    f"Registration too short: '{normalized}' (need at least 4 characters).\n"
                    "Ask: 'Could you give me the full registration?'"
                )
            
            # Validation: must contain at least one digit
            if not any(c.isdigit() for c in normalized):
                # Check if this is the caller's name echoed back
                first_up = self._state.customer_name_first.upper()
                last_up = self._state.customer_name_last.upper()
                if normalized == first_up or normalized == last_up or normalized == (first_up + last_up):
                    logger.info(f"[LOOKUP] Ignored caller name '{normalized}' passed as VRN (echo)")
                    return (
                        f"IGNORED: '{normalized}' is the caller's name, not a registration.\n"
                        "Ask: 'Could I grab your vehicle registration?'"
                    )
                
                return (
                    f"Invalid registration: '{normalized}' has no numbers.\n"
                    "UK registrations always contain digits.\n"
                    "Ask: 'Could I grab your registration?'"
                )
            
            # Validation: max length (UK VRMs are 7 chars max)
            if len(normalized) > 7:
                logger.warning(f"[LOOKUP] VRN too long: '{normalized}' ({len(normalized)} chars) — truncating to 7")
                normalized = normalized[:7]
            
            # Store for confirmation step
            self._state.vrn_pending = normalized
            
            # Generate phonetic readback
            phonetics = vrm_to_phonetics(normalized)
            
            logger.info(f"[LOOKUP] Parsed '{reg}' → '{normalized}' → phonetics: {phonetics}")
            
            return (
                f"Parsed registration: {normalized}\n"
                f"Phonetic readback: {phonetics}\n\n"
                f"Say to caller: '{phonetics}. Is that right?'\n\n"
                f"If YES → call lookup_vehicle(reg='{normalized}', confirmed=true)\n"
                f"If NO → Say: 'Let me get that again. Could you spell it out letter by letter?' "
                f"Then call lookup_vehicle with their new input."
            )

        @function_tool
        async def confirm_vehicle(
            context: RunContext,
            confirmed: bool,
            corrected_first_name: str = "",
            corrected_last_name: str = "",
        ) -> str:
            """Confirm or reject the vehicle lookup result. Pass corrected name if the caller corrected it."""

            if self._state.step != Step.CONFIRMING_VEHICLE:
                if self._state.step == Step.GREETING:
                    return (
                        "BLOCKED: You must call save_caller_name FIRST before any other tool. "
                        "Get the caller's name and intent, then call save_caller_name."
                    )
                return f"ERROR: Wrong step ({self._state.step.value}). No vehicle to confirm."

            if not confirmed:
                self._state.step = Step.NEED_VRN
                self._state.session_id = ""
                self._state.vehicle_make = ""
                self._state.vehicle_model = ""
                logger.info("[CONFIRM] Vehicle rejected → NEED_VRN")
                if self._state.vrn_attempts >= 3:
                    self._state.step = Step.MESSAGE_ONLY
                    return "Vehicle rejected after 3 attempts. Take their details for a callback."
                return "Vehicle rejected. Ask: 'No worries, could you read it out again for me?'"

            # Apply name corrections
            if corrected_first_name:
                self._state.customer_name_first = corrected_first_name.strip()
            if corrected_last_name:
                self._state.customer_name_last = corrected_last_name.strip()
            self._state.vrn_confirmed = True

            # Vehicle update path — skip service selection, go to take_message
            if self._state.intent == "vehicle_update":
                make = self._state.vehicle_make
                model = self._state.vehicle_model
                self._state.step = Step.MESSAGE_ONLY
                first = self._state.customer_name_first
                logger.info(f"[CONFIRM] Vehicle update path — {make} {model}")
                return (
                    f"Vehicle confirmed: {make.title()} {model.title()}.\n"
                    f"Say: 'Lovely, I've got that pulled up, {first}. What would you like me to pass on to the team?'\n"
                    "Collect their message and phone number, then call take_message.\n"
                    "Close: 'I'll make sure the team gets this. They'll give you a ring back with an update shortly.'"
                )

            # Prefetch services
            services: list[dict] = []
            if self._state.session_id:
                try:
                    all_services = await self._gh.list_services(self._state.session_id)
                    
                    # Filter services based on service type
                    if SERVICE_TYPE == "fast-fit":
                        # Fast-fit: include common services + full service, exclude major repairs
                        fast_fit_keywords = ["brake", "tyre", "mot", "oil", "battery", "wiper", "bulb", "air con", "full service", "service"]
                        exclude_keywords = ["major service", "clutch", "timing belt", "gearbox", "engine rebuild"]
                        
                        services = []
                        for svc in all_services:
                            svc_name = svc.get("name", "").lower()
                            # Exclude if contains any exclude keyword
                            if any(exclude in svc_name for exclude in exclude_keywords):
                                logger.info(f"[CONFIRM] Filtered out: {svc.get('name')}")
                                continue
                            # Include if contains any fast-fit keyword OR is "Other"
                            if any(keyword in svc_name for keyword in fast_fit_keywords) or "other" in svc_name:
                                services.append(svc)
                        
                        logger.info(f"[CONFIRM] Fast-fit mode: {len(services)} services (from {len(all_services)} total)")
                    else:
                        # Full-service: include everything
                        services = all_services
                        logger.info(f"[CONFIRM] Full-service mode: {len(services)} services")
                    
                    self._state.services_available = services
                except Exception as e:
                    logger.warning(f"[CONFIRM] Service prefetch failed: {e}")

            self._state.step = Step.NEED_SERVICE

            # Format service list
            svc_summary = ""
            if services:
                svc_lines = [
                    f"- {s.get('name', '?')} (ID: {s.get('service_price_id', '?')}{(', ' + _format_price(s)) if _format_price(s) else ''})"
                    for s in services[:8]
                ]
                svc_summary = "\nAvailable services:\n" + "\n".join(svc_lines)

            hint = self._state.service_hint
            if hint:
                matched = match_service(hint, services)
                if matched:
                    svc_name = matched.get('name')
                    price_str = _format_price(matched)
                    price_spoken = f" — {price_str}" if price_str else ""
                    return (
                        f"Vehicle confirmed.{svc_summary}\n\n"
                        f"The caller mentioned '{hint}'. Best match: {svc_name}{price_spoken}.\n"
                        f"Tell the caller: 'I'd suggest a {svc_name}{price_spoken} — shall I go with that?'\n"
                        f"If YES → call select_service(service_name='{svc_name}') with ZERO SPEECH.\n"
                        "If NO → ask what they'd prefer."
                    )
                # Service Advisor specialist: ask LLM to match the vague hint
                suggestion = await specialist_service_match(hint, services)
                if suggestion:
                    svc, reason = suggestion
                    svc_name = svc.get("name", "?")
                    
                    # Don't suggest "Other" to the caller - silently book it instead
                    is_other_category = 'other' in svc_name.lower() or 'general' in svc_name.lower()
                    if not is_other_category:
                        price_str = _format_price(svc)
                        price_spoken = f" ({price_str})" if price_str else ""
                        reason_line = f" — {reason}" if reason else ""
                        return (
                            f"Vehicle confirmed.{svc_summary}\n\n"
                            f"The caller mentioned '{hint}'{reason_line}.\n"
                            f"Suggest: '{svc_name}'{price_spoken}.\n"
                            f"Tell the caller: 'I'd suggest a {svc_name}{price_spoken} — shall I go with that?'\n"
                            f"If YES → call select_service(service_name='{svc_name}') with ZERO SPEECH.\n"
                            "If NO → ask what they'd prefer."
                        )
                    
                    # If specialist matched "Other" - ask for more detail, then book as "Other"
                    logger.info(f"[CONFIRM_VEHICLE] Specialist matched '{svc_name}' (Other category) - not suggesting, asking for service instead")
                    return (
                        f"Vehicle confirmed.{svc_summary}\n\n"
                        f"The caller mentioned '{hint}' — this will be booked under a general slot. "
                        f"Ask ONE short follow-up question to clarify the work needed (e.g. 'What work does it need on the {hint}?'). "
                        f"Once you have their answer, call select_service(service_name='{svc_name}') immediately with ZERO SPEECH. "
                        f"Do NOT try to match a different service from the list — use '{svc_name}' exactly."
                    )
                return (
                    f"Vehicle confirmed.{svc_summary}\n\n"
                    f"The caller mentioned '{hint}' but no exact match. "
                    "Ask which service they'd like, then call select_service with exactly what they describe."
                )

            return (
                f"Vehicle confirmed.{svc_summary}\n\n"
                "Ask: 'What work does it need?' Wait for their answer, then call select_service."
            )

        @function_tool
        async def select_service(context: RunContext, service_name: str) -> str:
            """Select a service by name. Fuzzy matching is applied automatically."""

            if self._state.step != Step.NEED_SERVICE:
                if self._state.step == Step.GREETING:
                    return (
                        "BLOCKED: You must call save_caller_name FIRST before any other tool. "
                        "Get the caller's name and intent, then call save_caller_name."
                    )
                return f"ERROR: Wrong step ({self._state.step.value}). Service selection not needed now."

            # Check if caller is requesting a major repair in fast-fit mode
            if SERVICE_TYPE == "fast-fit":
                major_repair_keywords = [
                    "engine replacement", "engine rebuild", "new engine",
                    "clutch", "clutch replacement", "new clutch",
                    "gearbox", "gearbox replacement", "transmission",
                    "timing belt", "timing chain", "cam belt",
                    "major service", "major repair",
                    "turbo", "turbocharger",
                    "head gasket"
                ]
                service_lower = service_name.lower()
                if any(keyword in service_lower for keyword in major_repair_keywords):
                    logger.info(f"[SELECT_SERVICE] Major repair requested in fast-fit mode: {service_name}")
                    self._state.step = Step.MESSAGE_ONLY
                    return (
                        f"MAJOR REPAIR REQUEST IN FAST-FIT MODE: '{service_name}'\n\n"
                        "Say naturally: 'As this booking's a little more in-depth, I'll need to take your details "
                        "and get one of the team to give you a ring back to arrange that.'\n"
                        "Then call take_message() to collect their details."
                    )

            services = self._state.services_available
            if not services and self._state.session_id:
                try:
                    services = await self._gh.list_services(self._state.session_id)
                    self._state.services_available = services
                except Exception as e:
                    logger.error(f"[SELECT_SERVICE] Failed to fetch services: {e}")
                    self._state.step = Step.MESSAGE_ONLY
                    return "Failed to load services. Take their details for a callback."

            if not services:
                self._state.step = Step.MESSAGE_ONLY
                return "No services available for this vehicle. Take their details for a callback."

            matched = match_service(service_name, services)
            if not matched:
                # Check if this is a diagnostic/symptom description
                diagnostic_result = await specialist_diagnostic_questions(service_name)
                if diagnostic_result:
                    questions = diagnostic_result.get("questions", [])
                    symptom_type = diagnostic_result.get("symptom_type", "unknown")
                    
                    # Store the symptom description for notes
                    if not self._state.diagnostic_notes:
                        self._state.diagnostic_notes = []
                    self._state.diagnostic_notes.append(f"Initial symptom: {service_name}")
                    self._state.diagnostic_notes.append(f"Symptom type: {symptom_type}")
                    
                    # Format questions for the agent
                    questions_formatted = "\n".join(f"{i+1}. {q}" for i, q in enumerate(questions))
                    logger.info(f"[DIAGNOSTIC] Starting {symptom_type} questionnaire for: {service_name}")
                    
                    return (
                        f"DIAGNOSTIC MODE: Customer described a {symptom_type} issue.\n"
                        f"Initial description: '{service_name}'\n\n"
                        f"IMPORTANT: Follow the structured diagnostic flow. Ask these questions ONE AT A TIME:\n"
                        f"{questions_formatted}\n\n"
                        f"After each answer, use record_diagnostic_info to save the response.\n"
                        f"Once all questions are answered, say: 'Right, based on what you've told me, "
                        f"I'd recommend a Diagnostic Check to identify the exact issue — shall I book that in?'\n"
                        "Then call select_service(service_name='Diagnostic Check') with ZERO SPEECH."
                    )
                
                # Service Advisor specialist: ask LLM to match the vague description
                suggestion = await specialist_service_match(service_name, services)
                if suggestion:
                    svc, reason = suggestion
                    svc_name = svc.get("name", "?")
                    
                    # Don't suggest "Other" to the caller - silently book it instead
                    is_other_category = 'other' in svc_name.lower() or 'general' in svc_name.lower()
                    if not is_other_category:
                        price_str = _format_price(svc)
                        price_spoken = f" ({price_str})" if price_str else ""
                        reason_line = f" — {reason}" if reason else ""
                        return (
                            f"No exact service name, but based on what the caller said{reason_line}.\n"
                            f"Suggest: '{svc_name}'{price_spoken}.\n"
                            f"Tell the caller: 'I'd suggest a {svc_name}{price_spoken} — shall I go with that?'\n"
                            "If YES → call select_service(service_name='" + svc_name + "') with ZERO SPEECH.\n"
                            "If NO → ask what they'd prefer."
                        )
                    
                    # If specialist matched "Other" category, book it silently without mentioning the category
                    logger.info(f"[SELECT_SERVICE] Specialist matched '{svc_name}' (Other category) - booking silently for: {service_name}")
                    svc_id = str(svc.get("service_price_id", ""))
                    
                    # Check if original service request is tyre-related before booking
                    is_tyre_request = any(keyword in service_name.lower() for keyword in ['tyre', 'tire'])
                    if is_tyre_request:
                        # Check if we need tyre position
                        if not self._state.tyre_position:
                            return (
                                f"Matched service: {service_name}.\n"
                                "TYRE POSITION REQUIRED.\n"
                                "Ask the caller: 'Which tyres need replacing?'\n"
                                "Wait for their answer, then call collect_tyre_info(tyre_position='their answer')."
                            )
                        
                        # Check if we need tyre size
                        if not self._state.tyre_size:
                            return (
                                f"Matched service: {service_name}.\n"
                                f"Tyre position: {self._state.tyre_position}\n"
                                "TYRE SIZE REQUIRED.\n"
                                "Ask the caller: 'What size tyres do you need? You can usually find this on the sidewall of your tyre.'\n"
                                "Wait for their answer, then call collect_tyre_info(tyre_size='their answer')."
                            )
                        
                        # Check if we need tyre quality
                        if not self._state.tyre_quality:
                            return (
                                f"Matched service: {service_name}.\n"
                                f"Tyre position: {self._state.tyre_position}\n"
                                f"Tyre size: {self._state.tyre_size}\n"
                                "TYRE QUALITY REQUIRED.\n"
                                "Ask the caller: 'Are you looking for budget, mid-range, or premium tyres?'\n"
                                "Wait for their answer, then call collect_tyre_info(tyre_quality='their answer')."
                            )
                    
                    try:
                        await self._gh.set_service(self._state.session_id, svc_id)
                        self._state.service_selected_id = svc_id
                        self._state.service_selected_name = svc_name
                        self._state.service_price = _format_price(svc)
                        self._state.other_service_description = service_name  # store original description for notes
                        logger.info(f"[SELECT_SERVICE] Booked under '{svc_name}' for: {service_name}")
                        
                        # Prefetch timeslots
                        timeslots: list[dict] = []
                        try:
                            timeslots = await self._gh.list_timeslots(self._state.session_id)
                            self._state.timeslots_available = timeslots
                        except Exception as e:
                            logger.warning(f"[SELECT_SERVICE] Timeslot prefetch failed: {e}")
                        
                        self._state.step = Step.NEED_TIMESLOT
                        
                        # If no timeslots available, take a message for callback
                        if not timeslots:
                            logger.warning(f"[SELECT_SERVICE] No timeslots available for {service_name}")
                            self._state.step = Step.MESSAGE_ONLY
                            return (
                                f"Service selected: {service_name}.\n"
                                "NO TIMESLOTS AVAILABLE.\n\n"
                                "Say naturally: 'I don't have any available slots at the moment. Can I take your number and we'll give you a call back to arrange a time?'\n"
                                "Then call take_message() to collect their details."
                            )
                        
                        slot_summary = ""
                        if timeslots:
                            slot_lines = [f"- {s['date']} at {s['time']}" for s in timeslots[:9]]
                            extra = f" (+{len(timeslots)-9} more available)" if len(timeslots) > 9 else ""
                            slot_summary = f"\nAvailable timeslots (showing {min(9, len(timeslots))} of {len(timeslots)}{extra}):\n" + "\n".join(slot_lines)
                        
                        return (
                            f"Service selected: {service_name}.{slot_summary}\n\n"
                            "Say naturally: 'I can book that in for you. When would suit you?'\n"
                            "Offer 2-3 early timeslots from the list above. If the caller asks for a date not listed, call select_timeslot with their preference — more slots exist beyond the ones shown."
                        )
                    except Exception as e:
                        logger.error(f"[SELECT_SERVICE] Failed to set Other service: {e}")
                        return f"Failed to set service: {e}. Try again."
                
                # No match - book under "Other" or general category (DON'T tell customer there's no option)
                logger.warning(f"[SELECT_SERVICE] No match for '{service_name}' - booking under Other")
                other_service = None
                for svc in services:
                    name_lower = svc.get('name', '').lower()
                    if 'other' in name_lower or 'general' in name_lower:
                        other_service = svc
                        break
                
                if other_service:
                    svc_id = str(other_service.get("service_price_id", ""))
                    svc_name = other_service.get("name", "Other")
                    
                    # Check if original service request is tyre-related
                    is_tyre_request = any(keyword in service_name.lower() for keyword in ['tyre', 'tire'])
                    if is_tyre_request:
                        # Check if we need tyre position
                        if not self._state.tyre_position:
                            return (
                                f"Matched service: {service_name}.\n"
                                "TYRE POSITION REQUIRED.\n"
                                "Ask the caller: 'Which tyres need replacing?'\n"
                                "Wait for their answer, then call collect_tyre_info(tyre_position='their answer')."
                            )
                        
                        # Check if we need tyre size
                        if not self._state.tyre_size:
                            return (
                                f"Matched service: {service_name}.\n"
                                f"Tyre position: {self._state.tyre_position}\n"
                                "TYRE SIZE REQUIRED.\n"
                                "Ask the caller: 'What size tyres do you need? You can usually find this on the sidewall of your tyre.'\n"
                                "Wait for their answer, then call collect_tyre_info(tyre_size='their answer')."
                            )
                        
                        # Check if we need tyre quality
                        if not self._state.tyre_quality:
                            return (
                                f"Matched service: {service_name}.\n"
                                f"Tyre position: {self._state.tyre_position}\n"
                                f"Tyre size: {self._state.tyre_size}\n"
                                "TYRE QUALITY REQUIRED.\n"
                                "Ask the caller: 'Are you looking for budget, mid-range, or premium tyres?'\n"
                                "Wait for their answer, then call collect_tyre_info(tyre_quality='their answer')."
                            )
                    
                    # Call set_service directly and continue
                    try:
                        await self._gh.set_service(self._state.session_id, svc_id)
                        self._state.service_selected_id = svc_id
                        self._state.service_selected_name = svc_name
                        self._state.service_price = _format_price(other_service)
                        logger.info(f"[SELECT_SERVICE] Booked under '{svc_name}' for: {service_name}")
                        
                        # Prefetch timeslots
                        timeslots: list[dict] = []
                        try:
                            timeslots = await self._gh.list_timeslots(self._state.session_id)
                            self._state.timeslots_available = timeslots
                        except Exception as e:
                            logger.warning(f"[SELECT_SERVICE] Timeslot prefetch failed: {e}")
                        
                        self._state.step = Step.NEED_TIMESLOT
                        
                        # If no timeslots available, take a message for callback
                        if not timeslots:
                            logger.warning(f"[SELECT_SERVICE] No timeslots available for {service_name}")
                            self._state.step = Step.MESSAGE_ONLY
                            return (
                                f"Service selected: {service_name}.\n"
                                "NO TIMESLOTS AVAILABLE.\n\n"
                                "Say naturally: 'I don't have any available slots at the moment. Can I take your number and we'll give you a call back to arrange a time?'\n"
                                "Then call take_message() to collect their details."
                            )
                        
                        slot_summary = ""
                        first_slot = ""
                        if timeslots:
                            slot_lines = [f"- {s['date']} at {s['time']}" for s in timeslots[:9]]
                            extra = f" (+{len(timeslots)-9} more available)" if len(timeslots) > 9 else ""
                            slot_summary = f"\nAvailable timeslots (showing {min(9, len(timeslots))} of {len(timeslots)}{extra}):\n" + "\n".join(slot_lines)
                            first_slot = f"{timeslots[0]['date']} at {timeslots[0]['time']}"
                        
                        return (
                            f"Service selected: {service_name}.{slot_summary}\n\n"
                            f"Say naturally: 'The next available slot is {first_slot}, or do you have a date in mind?'\n"
                            "Wait for their preference, then call select_timeslot. If the caller asks for a later date, call select_timeslot with their preference — more slots exist beyond those shown."
                        )
                    except Exception as e:
                        logger.error(f"[SELECT_SERVICE] Failed to set Other service: {e}")
                        return f"Failed to set service: {e}. Try again."
                
                # Fallback if no "Other" service exists - list available services
                svc_lines = [f"- {s.get('name', '?')}" for s in services[:6]]
                return (
                    f"No match for '{service_name}'. Available:\n" + "\n".join(svc_lines) + "\n"
                    "Ask the caller to choose from these."
                )

            svc_id = str(matched.get("service_price_id", ""))
            svc_name = matched.get("name", service_name)
            price = _format_price(matched)
            
            # TYRE BOOKING: If service is tyre-related, collect position, size and quality first
            is_tyre_service = any(keyword in svc_name.lower() for keyword in ['tyre', 'tire'])
            if is_tyre_service:
                # Check if we need tyre position
                if not self._state.tyre_position:
                    return (
                        f"Matched service: {svc_name}.\n"
                        "TYRE POSITION REQUIRED.\n"
                        "Ask the caller: 'Which tyres need replacing?'\n"
                        "Wait for their answer, then call collect_tyre_info(tyre_position='their answer')."
                    )
                
                # Check if we need tyre size
                if not self._state.tyre_size:
                    return (
                        f"Matched service: {svc_name}.\n"
                        f"Tyre position: {self._state.tyre_position}\n"
                        "TYRE SIZE REQUIRED.\n"
                        "Ask the caller: 'What size tyres do you need? You can usually find this on the sidewall of your tyre.'\n"
                        "Wait for their answer, then call collect_tyre_info(tyre_size='their answer')."
                    )
                
                # Check if we need tyre quality
                if not self._state.tyre_quality:
                    return (
                        f"Matched service: {svc_name}.\n"
                        f"Tyre position: {self._state.tyre_position}\n"
                        f"Tyre size: {self._state.tyre_size}\n"
                        "TYRE QUALITY REQUIRED.\n"
                        "Ask the caller: 'Are you looking for budget, mid-range, or premium tyres?'\n"
                        "Wait for their answer, then call collect_tyre_info(tyre_quality='their answer')."
                    )

            # Set service via API
            try:
                await self._gh.set_service(self._state.session_id, svc_id)
            except Exception as e:
                logger.error(f"[SELECT_SERVICE] API failed: {e}")
                return f"Failed to set service: {e}. Try again."

            self._state.service_selected_id = svc_id
            self._state.service_selected_name = svc_name
            self._state.service_price = price
            # If booked under "Other" category, store the original hint for GarageHive notes
            if 'other' in svc_name.lower() or 'general' in svc_name.lower():
                if self._state.service_hint and not self._state.other_service_description:
                    self._state.other_service_description = self._state.service_hint
            logger.info(f"[SELECT_SERVICE] Set: {svc_name} ({price or 'no price'})")

            # Prefetch timeslots
            timeslots: list[dict] = []
            try:
                timeslots = await self._gh.list_timeslots(self._state.session_id)
                self._state.timeslots_available = timeslots
                logger.info(f"[SELECT_SERVICE] Prefetched {len(timeslots)} timeslots")
            except Exception as e:
                logger.warning(f"[SELECT_SERVICE] Timeslot prefetch failed: {e}")

            slot_summary = ""
            first_slot = ""
            if timeslots:
                slot_lines = [f"- {s['date']} at {s['time']}" for s in timeslots[:9]]
                extra = f" (+{len(timeslots)-9} more available)" if len(timeslots) > 9 else ""
                slot_summary = f"\nAvailable timeslots (showing {min(9, len(timeslots))} of {len(timeslots)}{extra}):\n" + "\n".join(slot_lines)
                first_slot = f"{timeslots[0]['date']} at {timeslots[0]['time']}"

            self._state.step = Step.NEED_TIMESLOT
            vehicle_desc = f"{self._state.vehicle_make.title()} {self._state.vehicle_model.title()}".strip()

            # Quote flow
            if self._state.intent == "quote":
                price_str = price if price else "available on request"
                return (
                    f"Service set: {svc_name} ({price_str}).{slot_summary}\n\n"
                    f"NOW tell the caller: 'A {svc_name} for your {vehicle_desc} would be {price_str}.'\n"
                    "Then ask: 'Would you like me to book that in for you?'\n"
                    f"If YES → say 'The next available slot is {first_slot}, or do you have a date in mind?' and wait for their preference. If they ask for a date not in the list, call select_timeslot with their preference — more slots are available beyond those shown.\n"
                    "If NO → say 'No worries, I'll get one of the team to give you a ring if you change your mind.' "
                    "then call take_message."
                )

            # Booking flow
            price_str = f" — {price}" if price else ""
            return (
                f"Service set: {svc_name}{price_str}.{slot_summary}\n\n"
                f"Say naturally: 'The next available slot is {first_slot}, or do you have a date in mind?'\n"
                "Wait for their preference, then call select_timeslot. If the caller asks for a later date not in the list, call select_timeslot with their preference — more slots are available."
            )

        @function_tool
        async def record_diagnostic_info(context: RunContext, diagnostic_answer: str) -> str:
            """Record the caller's answer to a diagnostic question.
            Use this to save important details about symptoms, timing, frequency, etc.
            Example: 'When braking at speed', 'Constant for 2 weeks', 'Engine light came on yesterday'"""
            
            if not self._state.diagnostic_notes:
                self._state.diagnostic_notes = []
            
            self._state.diagnostic_notes.append(diagnostic_answer)
            logger.info(f"[DIAGNOSTIC] Recorded: {diagnostic_answer}")
            
            return (
                f"Diagnostic info recorded: '{diagnostic_answer}'\n"
                "Continue with the next diagnostic question if there are more, "
                "or proceed to recommend a Diagnostic Check service."
            )

        @function_tool
        async def collect_tyre_info(
            context: RunContext,
            tyre_position: str = "",
            tyre_size: str = "",
            tyre_quality: str = ""
        ) -> str:
            """Collect tyre position, size and quality preference for tyre bookings.
            tyre_position: e.g., 'front left', 'both fronts', 'all four', 'rear right'
            tyre_size: e.g., '205/55 R16', '225/45/17'
            tyre_quality: 'budget', 'mid-range', or 'premium'"""
            
            if tyre_position:
                self._state.tyre_position = tyre_position.strip()
                logger.info(f"[TYRE_INFO] Position: {self._state.tyre_position}")
            
            if tyre_size:
                self._state.tyre_size = tyre_size.strip()
                logger.info(f"[TYRE_INFO] Size: {self._state.tyre_size}")
            
            if tyre_quality:
                self._state.tyre_quality = tyre_quality.strip().lower()
                logger.info(f"[TYRE_INFO] Quality: {self._state.tyre_quality}")
            
            # Check if we have all three pieces of info
            if self._state.tyre_position and self._state.tyre_size and self._state.tyre_quality:
                return (
                    f"Tyre info collected: {self._state.tyre_position}, Size {self._state.tyre_size}, {self._state.tyre_quality} quality.\n"
                    "Now call select_service again with the same service name to continue booking."
                )
            elif self._state.tyre_position and self._state.tyre_size and not self._state.tyre_quality:
                return (
                    f"Tyre position and size recorded: {self._state.tyre_position}, {self._state.tyre_size}.\n"
                    "NEXT: Ask the caller about tyre quality.\n"
                    "Say: 'Are you looking for budget, mid-range, or premium tyres?'\n"
                    "Then call collect_tyre_info(tyre_quality='their answer')."
                )
            elif self._state.tyre_position and not self._state.tyre_size:
                return (
                    f"Tyre position recorded: {self._state.tyre_position}.\n"
                    "Now ask: 'What size tyres do you need? You can usually find this on the sidewall of your tyre.'\n"
                    "Then call collect_tyre_info(tyre_size='their answer')."
                )
            elif not self._state.tyre_position:
                return (
                    "Tyre position not set yet.\n"
                    "Ask: 'Which tyres need replacing?'\n"
                    "Then call collect_tyre_info(tyre_position='their answer')."
                )
            else:
                return "Please provide tyre_position, tyre_size, or tyre_quality parameter."

        @function_tool
        async def select_timeslot(context: RunContext, caller_preference: str) -> str:
            """Set the chosen timeslot. Pass the caller's words about when they want to come in.
            Examples: 'Thursday morning', 'the 9:30 one', 'tomorrow at 2pm', 'as soon as possible'.
            The tool handles all date/time parsing against available slots."""

            if self._state.step != Step.NEED_TIMESLOT:
                if self._state.step == Step.GREETING:
                    return (
                        "BLOCKED: You must call save_caller_name FIRST before any other tool. "
                        "Get the caller's name and intent, then call save_caller_name."
                    )
                return f"ERROR: Wrong step ({self._state.step.value}). Timeslot selection not needed now."
            
            # Validate service was selected before allowing timeslot booking
            if not self._state.service_selected_name:
                return (
                    "BLOCKED: No service has been selected yet. "
                    "You must call select_service to choose what work needs doing before booking a timeslot. "
                    "Go back and ask what service they need, then call select_service."
                )

            today_str = _current_uk_datetime().strftime("%A %d %B %Y (%Y-%m-%d)")

            # Try specialist LLM first
            match = await specialist_timeslot_match(
                caller_preference, self._state.timeslots_available, today_str
            )

            # Fallback: try regex extraction of YYYY-MM-DD and HH:MM
            # (backward compat if LLM passes formatted dates)
            if match is None:
                date_re = re.search(r"(\d{4}-\d{2}-\d{2})", caller_preference)
                time_re = re.search(r"(\d{1,2}:\d{2})", caller_preference)
                if date_re and time_re:
                    match = {"date": date_re.group(1), "time": time_re.group(1)}
                    logger.info(f"[SELECT_TIMESLOT] Regex fallback: {match}")

            if match is None:
                # No match — list available slots
                slots = self._state.timeslots_available
                if slots:
                    # Group slots by date to show range more clearly
                    dates_seen = {}
                    for s in slots:
                        d = s['date']
                        if d not in dates_seen:
                            dates_seen[d] = s['time']
                    # Show first 9 slots (covers ~3 days)
                    slot_lines = [f"- {s['date']} at {s['time']}" for s in slots[:9]]
                    first_date = slots[0]['date'] if slots else '?'
                    last_date = slots[-1]['date'] if slots else '?'
                    return (
                        f"Couldn't match '{caller_preference}' to an available slot.\n"
                        f"Slots available from {first_date} to {last_date}:\n"
                        "Available timeslots:\n" + "\n".join(slot_lines) + "\n"
                        "Read 2-3 options from different days and ask: 'Which works best for you?'"
                    )
                return "No timeslots available. Take their details for a callback."

            booking_date = match["date"]
            booking_time = match["time"]

            try:
                await self._gh.set_timeslot(self._state.session_id, booking_date, booking_time)
            except Exception as e:
                logger.error(f"[SELECT_TIMESLOT] API failed: {e}")
                return f"Failed to set timeslot: {e}. Try a different slot."

            self._state.booking_date = booking_date.strip()
            self._state.booking_time = booking_time.strip()
            self._state.step = Step.NEED_CONTACT
            logger.info(f"[SELECT_TIMESLOT] Set: {booking_date} at {booking_time}")

            surname_note = ""
            if not self._state.customer_name_last:
                surname_note = "IMPORTANT: No surname on file yet. Ask 'And your surname?' first.\n"

            return (
                f"Timeslot set: {booking_date} at {booking_time}.\n"
                f"{surname_note}"
                "Say: 'Lovely. I just need a couple of details.' then ask for their surname if missing, "
                "otherwise ask: 'What's the best number for you?'\n"
                "When they give their surname, call update_caller_name(last_name='...') to save it, "
                "then KEEP addressing them by their FIRST name — not the surname.\n"
                "Collect ONE field at a time: surname → phone → email → postcode (call validate_address) → house number.\n"
                "You MUST collect ALL five fields AND call validate_address BEFORE calling submit_booking.\n"
                "Do NOT call submit_booking until you have phone, email, postcode, AND house number."
            )

        @function_tool
        async def validate_address(context: RunContext, postcode: str) -> str:
            """Validate a UK postcode and look up the area. Call AFTER collecting the postcode from the caller."""

            if self._state.step != Step.NEED_CONTACT:
                if self._state.step == Step.GREETING:
                    return (
                        "BLOCKED: You must call save_caller_name FIRST before any other tool. "
                        "Get the caller's name and intent, then call save_caller_name."
                    )
                return f"ERROR: Wrong step ({self._state.step.value}). Address validation not needed now."

            # Sanity check: postcode must be real, not "n/a" or empty
            clean_pc = (postcode or "").strip().lower()
            if not clean_pc or clean_pc in ("n/a", "na", "none", "no"):
                return (
                    "No postcode provided. Ask the caller: 'And your postcode?' "
                    "UK postcodes are like 'SW1A 1AA' or 'M1 1AA'."
                )

            # Detect email address or domain passed as postcode
            _email_domains = ("gmail", "hotmail", "yahoo", "outlook", "icloud", "aol", "protonmail", "mail")
            if "@" in clean_pc or any(d in clean_pc for d in _email_domains) or ".com" in clean_pc or ".co.uk" in clean_pc:
                return (
                    f"WRONG FIELD: '{postcode}' is an email address/domain, NOT a postcode. "
                    "You have NOT asked for the postcode yet — do that now. "
                    "Ask: 'And your postcode?' Do NOT call validate_address until they give an actual postcode."
                )

            # Detect phone number passed as postcode (digits with + prefix, or 7+ digits)
            digits_only = re.sub(r"[^0-9]", "", clean_pc)
            if clean_pc.startswith("+") or len(digits_only) >= 7:
                return (
                    f"WRONG FIELD: '{postcode}' is a PHONE NUMBER, NOT a postcode. "
                    "Store it for submit_booking later. "
                    "Now ask for the POSTCODE: 'And your postcode?' "
                    "Do NOT call validate_address until they give an actual UK postcode."
                )

            # UK postcode format check: must have letters AND digits, 5-8 chars
            pc_no_space = re.sub(r"\s", "", clean_pc)
            if (pc_no_space.isdigit()
                    or pc_no_space.isalpha()
                    or len(pc_no_space) < 5
                    or len(pc_no_space) > 8):
                return (
                    f"'{postcode}' doesn't look like a valid UK postcode. "
                    "UK postcodes have letters AND digits (e.g. 'SW1A 1AA', 'M1 1AA', 'B2 4QA'). "
                    "Ask: 'Sorry, could I get your postcode? It'd be something like SW1A 1AA.'"
                )

            result = await self._gh.validate_address(postcode)
            street = result.get("street", "")
            city = result.get("city", "")

            self._state.postcode = postcode.strip()
            self._state.street = street
            self._state.city = city
            logger.info(f"[VALIDATE_ADDRESS] {street}, {city}, {postcode}")

            if street and city:
                return (
                    f"Postcode found: {street}, {city}.\n"
                    f"Confirm with caller: 'Is that {street}, {city}?'\n"
                    "Once confirmed, ask: 'And the house number or name?'\n"
                    "After they give it, call submit_booking with ALL details."
                )
            if city:
                return (
                    f"Postcode found: {city}.\n"
                    f"Confirm with caller: 'Is that the {city} area?'\n"
                    "Once confirmed, ask: 'And the house number or name?'\n"
                    "After they give it, call submit_booking with ALL details. "
                    "No street name needed — some postcodes don't have one."
                )
            return (
                f"Postcode accepted: {postcode} (area details not found).\n"
                "Now ask: 'And the house number or name?'\n"
                "After they give it, call submit_booking with ALL details."
            )

        @function_tool
        async def submit_booking(
            context: RunContext,
            phone: str,
            email: str,
            house_name_or_number: str,
            postcode: str,
            street: str = "",
            city: str = "",
            notes: str = "",
        ) -> str:
            """Finalize the booking. ALL fields except notes are required."""

            if self._state.step != Step.NEED_CONTACT:
                if self._state.step == Step.GREETING:
                    return (
                        "BLOCKED: You must call save_caller_name FIRST before any other tool. "
                        "Get the caller's name and intent, then call save_caller_name."
                    )
                return f"ERROR: Wrong step ({self._state.step.value}). Cannot submit booking now."

            first = self._state.customer_name_first
            last = self._state.customer_name_last
            phone = _sanitise_phone((phone or "").strip())
            email = _sanitise_email((email or "").strip().replace(" ", "").lower())
            house_name_or_number = (house_name_or_number or "").strip()
            postcode = (postcode or "").strip()
            street = (street or self._state.street or "").strip()
            city = (city or self._state.city or "").strip()

            # Reject persona name
            if first.lower() == "leah" or last.lower() == "leah":
                return (
                    "ERROR: You passed your persona name 'Leah' as the contact name. "
                    "Use the CALLER's name from earlier in the conversation."
                )

            # Hard guard: verify booking pipeline completed (catches corrupted state)
            pipeline_missing = []
            if not self._state.session_id:
                pipeline_missing.append("booking session (lookup_vehicle never succeeded)")
            if not self._state.service_selected_name:
                pipeline_missing.append("service (select_service never called)")
            if not self._state.booking_date or not self._state.booking_time:
                pipeline_missing.append("timeslot (select_timeslot never called)")
            if pipeline_missing:
                return (
                    f"BLOCKED — booking pipeline incomplete: {', '.join(pipeline_missing)}. "
                    "These steps must complete before submitting. Go back and complete them first."
                )

            # Validate contact fields
            missing = []
            if not first:
                missing.append("first name")
            if not last:
                missing.append("surname")
            if not phone:
                missing.append("phone number")
            # Reject obviously fake/placeholder emails the LLM may hallucinate
            _fake_email_domains = ("domain.com", "example.com", "email.com", "test.com", "placeholder.com")
            if not email or "@" not in email:
                missing.append("valid email address")
            elif any(email.endswith(d) for d in _fake_email_domains):
                missing.append("valid email address ('" + email + "' looks fabricated — ask the caller for their real email)")
            if not house_name_or_number:
                missing.append("house name or number")
            if not postcode:
                missing.append("postcode")
            # Street and city come from validate_address — street can be legitimately empty
            # for some postcodes (e.g. SW1A 1AA). Only require validate_address was called.
            if not city and not self._state.city:
                missing.append("city (call validate_address first to look up the postcode)")
            if missing:
                self._state.booking_submit_pending = True
                # Give explicit fix instructions for name fields
                if not last or not first:
                    return (
                        f"Cannot submit — missing: {', '.join(missing)}.\n"
                        "Ask the caller for the missing name info, then call "
                        "update_caller_name(first_name='...', last_name='...') to save it. "
                        "After that, call submit_booking again."
                    )
            # Build notes: start with any agent-passed notes, then append extras
            all_notes = notes

            # Add service description for "Other" category bookings
            if self._state.other_service_description:
                all_notes = f"{self._state.other_service_description}\n\n{all_notes}".strip() if all_notes else self._state.other_service_description

            # Add tyre information if present
            if self._state.tyre_position or self._state.tyre_size or self._state.tyre_quality:
                tyre_info_parts = []
                if self._state.tyre_position:
                    tyre_info_parts.append(f"Position: {self._state.tyre_position}")
                if self._state.tyre_size:
                    tyre_info_parts.append(f"Tyre size: {self._state.tyre_size}")
                if self._state.tyre_quality:
                    tyre_info_parts.append(f"Quality: {self._state.tyre_quality}")
                tyre_section = "TYRE INFORMATION:\n" + "\n".join(tyre_info_parts)
                all_notes = f"{all_notes}\n\n{tyre_section}".strip() if all_notes else tyre_section

            # Add diagnostic notes if present
            if self._state.diagnostic_notes:
                diagnostic_section = "DIAGNOSTIC INFO:\n" + "\n".join(self._state.diagnostic_notes)
                all_notes = f"{all_notes}\n\n{diagnostic_section}".strip() if all_notes else diagnostic_section

            logger.info(f"[SUBMIT_BOOKING] Notes to GarageHive: '{all_notes}'")

            contact_address = f"{house_name_or_number}, {street}".strip(", ").lower()
            try:
                result = await self._gh.set_contact_info(
                    self._state.session_id,
                    contact_name=first,
                    contact_last_name=last,
                    contact_email=email,
                    contact_number=phone,
                    contact_address=contact_address,
                    contact_city=city.lower(),
                    contact_postcode=postcode,
                    contact_salutation=10,
                    contact_address2="",
                    notes=all_notes,
                )
            except Exception as e:
                self._state.booking_submit_pending = True
                logger.error(f"[SUBMIT_BOOKING] API error: {e}")
                asyncio.create_task(ErrorMonitor.report_error(
                    error_msg=f"submit_booking failed: {e}", agent_name="SUPERVISOR",
                    room_name=self._room_name, error_type="api_error",
                    extra={"session_id": self._state.session_id},
                ))
                return f"Booking API failed: {e}. Try again or take their details for a callback."

            if result.get("status") == "success":
                self._state.step = Step.CONFIRMED
                self._state.booking_submit_pending = False
                self._state.contact_phone = phone
                self._state.contact_email = email
                vehicle_desc = f"{self._state.vehicle_make.title()} {self._state.vehicle_model.title()}".strip()
                logger.info("[SUBMIT_BOOKING] CONFIRMED")

                # Format date naturally: "tomorrow", "Monday 18th February", etc.
                try:
                    bdate = datetime.strptime(self._state.booking_date, "%Y-%m-%d").date()
                    today = datetime.now(ZoneInfo("Europe/London")).date()
                    diff = (bdate - today).days
                    day_num = bdate.day
                    suffix = "th" if 11 <= day_num <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day_num % 10, "th")
                    if diff == 0:
                        natural_date = "today"
                    elif diff == 1:
                        natural_date = "tomorrow"
                    else:
                        natural_date = f"{bdate.strftime('%A')} the {day_num}{suffix} of {bdate.strftime('%B')}"
                except Exception:
                    natural_date = self._state.booking_date

                return (
                    "BOOKING CONFIRMED.\n"
                    f"Say: 'That's all booked in — {self._state.service_selected_name} "
                    f"for your {vehicle_desc} on {natural_date} at {self._state.booking_time}.'\n"
                    "Then: 'Anything else I can help with?'\n"
                    "When done: 'Brilliant, cheers then. Have a lovely day!'"
                )

            errors = result.get("errors", [])
            msg = result.get("message", "Unknown error")
            return f"Booking failed: {msg}. Errors: {errors}. Check the details and try again."

        @function_tool
        async def update_caller_name(
            context: RunContext,
            first_name: str = "",
            last_name: str = "",
        ) -> str:
            """Update the caller's name AFTER save_caller_name has already been called.
            Use when the caller corrects their name or gives their surname later."""

            if self._state.step == Step.GREETING:
                return (
                    "WRONG TOOL: You are still in GREETING step. "
                    "Call save_caller_name (not update_caller_name) to save the name AND set the intent. "
                    "save_caller_name advances the call. update_caller_name is only for corrections later."
                )

            updated = []
            if first_name and first_name.strip():
                self._state.customer_name_first = first_name.strip()
                updated.append(f"first_name='{first_name.strip()}'")
            if last_name and last_name.strip():
                self._state.customer_name_last = last_name.strip()
                updated.append(f"last_name='{last_name.strip()}'")
            if not updated:
                return "ERROR: Provide at least first_name or last_name to update."

            logger.info(f"[UPDATE_NAME] Updated: {', '.join(updated)}")

            # Give a specific next-action based on current step
            step_next = {
                Step.NEED_VRN: "Say ONE short sentence asking for their registration, then STOP and WAIT.",
                Step.CONFIRMING_VEHICLE: "Now confirm the vehicle with the caller.",
                Step.NEED_SERVICE: "Now ask what service they need.",
                Step.NEED_TIMESLOT: "Now offer available timeslots.",
                Step.NEED_CONTACT: "Now continue collecting contact details (phone → email → postcode → house number).",
            }
            next_action = step_next.get(self._state.step, "Continue the conversation.")

            return (
                f"Name updated: {', '.join(updated)}. "
                f"Full name on file: {self._state.customer_name_first} {self._state.customer_name_last}.\n"
                f"Keep addressing them as '{self._state.customer_name_first}' (first name only).\n"
                f"NEXT ACTION: {next_action}\n"
                "Do NOT call update_caller_name again unless the caller EXPLICITLY corrects their name."
            )

        @function_tool
        async def take_message(
            context: RunContext,
            message: str,
            phone: str,
            name_first: str = "",
            name_last: str = "",
            vrn: str = "",
            callback_time: str = "",
        ) -> str:
            """Save a message for the team to call back. Available as an escape hatch from any step."""

            allowed = (Step.GREETING, Step.MESSAGE_ONLY, Step.NEED_VRN, Step.NEED_SERVICE, Step.NEED_TIMESLOT, Step.NEED_CONTACT)
            if self._state.step not in allowed:
                return f"ERROR: Wrong step ({self._state.step.value}). Cannot take message now."

            self._state.message = (message or "").strip()
            self._state.contact_phone = (phone or "").strip()
            self._state.preferred_callback_time = (callback_time or "").strip()
            if name_first:
                self._state.customer_name_first = name_first.strip()
            if name_last:
                self._state.customer_name_last = name_last.strip()
            if vrn:
                self._state.vrn = normalize_vehicle_registration(vrn)

            self._state.step = Step.DONE
            first = self._state.customer_name_first
            last = self._state.customer_name_last
            logger.info(f"[TAKE_MESSAGE] From {first} {last}: {message}")

            # Specialist summariser — structure the message for the team
            vehicle_info = f"{self._state.vehicle_make} {self._state.vehicle_model}".strip()
            summary = await specialist_summarise_message(
                self._state.message, f"{first} {last}", vehicle_info
            )
            if summary:
                self._state.message_summary = summary
                logger.info(f"[TAKE_MESSAGE] Summary: {summary}")

            # Track escalation — include structured summary if available
            escalation_extra: dict[str, Any] = {
                "caller": f"{first} {last}",
                "intent": self._state.intent,
            }
            if summary:
                escalation_extra["category"] = summary.get("category", "")
                escalation_extra["summary"] = summary.get("summary", "")
                escalation_extra["action"] = summary.get("action", "")

            asyncio.create_task(ErrorMonitor.report_error(
                error_msg=f"Call escalated to message: {summary.get('summary', message) if summary else message}",
                agent_name="SUPERVISOR", room_name=self._room_name,
                error_type="escalation",
                extra=escalation_extra,
            ))

            return (
                f"Message saved from {first} {last}.\n"
                "Read back a brief summary to confirm.\n"
                "Then: 'Lovely, I'll make sure the team gets this. They'll give you a ring back shortly.'\n"
                "Close: 'Cheers, have a lovely day!'"
            )

        # ── System Prompt ────────────────────────────────────

        greeting = get_dynamic_greeting(AGENT_BRANCH_NAME)
        now = _current_uk_datetime()
        today_str = now.strftime("%A %d %B %Y")
        tomorrow = now.replace(hour=0, minute=0, second=0) + __import__("datetime").timedelta(days=1)
        tomorrow_str = tomorrow.strftime("%A %d %B %Y")

        # Different instructions based on mode
        if self._assist_mode:
            instructions = f"""YOU ARE LEAH — a warm, friendly British receptionist at {AGENT_BRANCH_NAME}.
One person, one voice, one natural conversation from start to finish.

TODAY: {today_str}. Tomorrow: {tomorrow_str} ({tomorrow.strftime("%Y-%m-%d")}).
OPENING HOURS: {get_business_hours_text()}

MODE: ASSIST MODE - You CANNOT make bookings. Your role is to help callers by:
- Answering questions about services, pricing, and opening hours
- Taking messages for the team to call back
- Collecting caller details for bookings that the team will process

PERSONALITY: Sound natural and warm, like a real person — not robotic. Vary your phrasing each turn.
- Mix short replies ("Brilliant.") with slightly longer ones ("Lovely, that's all popped in for you.")
- Use natural British phrases: "lovely", "brilliant", "no worries", "cheers", "pop it in", "give you a ring", "smashing"
- Say times naturally: 08:30 = "half eight in the morning", 14:00 = "two in the afternoon"
- NEVER use: "awesome", "gotten", "you guys", "super"
- Save "Cheers, have a lovely day!" for the very end of the call

PRONUNCIATION:
- Say "garage" as "garridge" (British pronunciation - rhymes with "carriage", NOT "ga-RAHJ")

RULES:
- NO filler before/between tool calls. No "one moment", "let me check", "bear with me".
- ONE QUESTION PER TURN. Ask one thing, then STOP.
- Use the caller's FIRST name sparingly (2-3 times max per call: greeting, key moments, closing).
- ALL prices are in British Pounds. Say "fifty pounds" not "fifty dollars".

FIRST STEP: save_caller_name. ALL tools LOCKED until it succeeds. Do NOT hallucinate names.

FLOW FOR ALL ENQUIRIES:
1. GREETING: "{greeting}" spoken. Get the caller's name first.
2. Understand what they need - booking, question, price check, callback request, etc.
3. For ANY booking request: Say naturally: "I can take all your details down and the team will give you a ring back to get that booked in. What work does your vehicle need?"
4. Collect: name, what they need done, vehicle registration (optional), phone number
5. Use take_message to save their details (pass null for callback_time)
6. CLOSE: "Lovely, I'll make sure the team gets this. They'll give you a ring back shortly. Cheers, have a lovely day!"

IMPORTANT: 
- You CANNOT access the booking system, check availability, or confirm appointments
- Do NOT ask when they'd like a callback - the team will call them back when available
- Always route to take_message for bookings"""
        else:
            instructions = f"""YOU ARE LEAH — a warm, friendly British receptionist at {AGENT_BRANCH_NAME}.
One person, one voice, one natural conversation from start to finish.

TODAY: {today_str}. Tomorrow: {tomorrow_str} ({tomorrow.strftime("%Y-%m-%d")}).
OPENING HOURS: {get_business_hours_text()}

PERSONALITY: Sound natural and warm, like a real person — not robotic. Vary your phrasing each turn.
- Mix short replies ("Brilliant.") with slightly longer ones ("Lovely, that's all popped in for you.")
- Use natural British phrases: "lovely", "brilliant", "no worries", "cheers", "pop it in", "give you a ring", "smashing"
- Say times naturally: 08:30 = "half eight in the morning", 14:00 = "two in the afternoon"
- NEVER use: "awesome", "gotten", "you guys", "super"
- Save "Cheers, have a lovely day!" for the very end of the call

PRONUNCIATION:
- Say "garage" as "garridge" (British pronunciation - rhymes with "carriage", NOT "ga-RAHJ")

RULES:
- NO filler before/between tool calls. No "one moment", "let me check", "bear with me".
- NEVER say "confirmed"/"booked" unless a tool directive tells you to.
- ONE QUESTION PER TURN. Ask one thing, then STOP.
- Each tool returns a directive — FOLLOW IT EXACTLY.
- Use the caller's FIRST name sparingly (2-3 times max per call: greeting, key moments, closing). DO NOT repeat their name after every sentence.
- ALL prices are in British Pounds. Say "fifty pounds" not "fifty dollars".
- If submit_booking returns an error (missing fields), you MUST collect the missing info and retry BEFORE ending the call. Do NOT say goodbye until the booking is confirmed or the caller explicitly gives up.

FIRST STEP: save_caller_name. ALL tools LOCKED until it succeeds. Do NOT hallucinate names.

INTENT DETECTION:
- "Can I speak to [name]" / "Is [name] available" / "Can I talk to a human" → intent='transfer', requested_person='[name]'
- "I dropped my car off" / "Checking on my vehicle" → intent='vehicle_update'
- "Just have a question" / "Need to reschedule" / "Want to cancel" → intent='message'
- "How much is a..." / "What's the price for..." → intent='quote'
- Default → intent='booking'

TRANSFER REQUESTS (caller wants to speak to a human/specific person):
- The tools will check business hours automatically and provide the correct response
- DURING business hours: Team is BUSY (not outside hours) - offer booking help or take a message
- OUTSIDE business hours: Team unavailable due to closure - take message for callback
- NEVER say "outside opening hours" when the garage is currently open

FLOW:
1. GREETING: "{greeting}" spoken. Get name + intent (booking/quote/vehicle update/message/transfer). Default booking. If no name: "Can I take your name?" then STOP.
   - If TRANSFER REQUEST: Follow the tool's instructions exactly - it will tell you whether the garage is open/closed
   - If VEHICLE UPDATE: Follow the tool's instructions exactly - it will tell you whether the garage is open/closed
2. VEHICLE: Call lookup_vehicle with caller's EXACT words. Tool handles NATO phonetics and gives you a phonetic readback to confirm with the caller.
3. SERVICE: Call select_service(service_name). Tool handles matching — just pass what the caller said.
   - DIAGNOSTIC FLOW: If the caller describes a fault/symptom (noise, warning light, problem), the tool will provide a structured diagnostic questionnaire:
     * STEP 1: Broad open question ("Can you tell me what it's doing?")
     * STEP 2: Clarify symptom type (when, how, circumstances)
     * STEP 3: Timing questions ("When did this first start?", "Has it got worse?")
   - Ask questions ONE AT A TIME. Wait for answer. Call record_diagnostic_info to save each response.
   - DO NOT interrupt. Let them speak fully.
   - After completing the questionnaire, recommend a Diagnostic Check.
4. TIMESLOT: Offer 2-3 early slots naturally. Call select_timeslot(caller_preference) with the caller's words — tool handles date parsing.
5. CONTACT (one at a time): surname → phone (read back last 3 digits) → email → postcode (call validate_address) → house number. Then submit_booking.
6. CLOSE: Confirm booking, "Cheers, have a lovely day!"

SPECIAL SITUATIONS - FOLLOW TOOL INSTRUCTIONS EXACTLY:
- TRANSFER REQUEST: The tool will check business hours and provide the exact script to use. DO NOT improvise.
  * During hours → "team are busy helping other customers" (NOT "outside hours")
  * Outside hours → "team unavailable outside opening hours"
- VEHICLE UPDATE: The tool will check business hours and provide the exact script. Follow it word-for-word.
- MESSAGE: Collect message → phone → callback time → take_message.
- CHANGE OF MIND: Booking↔Message works both ways.

CRITICAL: When the tool says "Say naturally: [exact phrase]", use that phrase. Don't mix phrases from different scenarios."""

        # Build tool list based on mode
        if self._assist_mode:
            # ASSIST MODE: Only basic tools, no GarageHive integration
            tool_list = [
                get_current_datetime,
                save_caller_name,
                update_caller_name,
                take_message,
            ]
        else:
            # AUTOMATE MODE: Full booking capability with GarageHive
            tool_list = [
                get_current_datetime,
                save_caller_name,
                update_caller_name,
                lookup_vehicle,
                confirm_vehicle,
                select_service,
                record_diagnostic_info,
                collect_tyre_info,
                select_timeslot,
                validate_address,
                submit_booking,
                take_message,
            ]

        super().__init__(
            instructions=instructions,
            tools=tool_list,
        )

    def set_session(self, session: AgentSession) -> None:
        self._agent_session = session

    async def on_enter(self) -> None:
        if self._agent_session:
            greeting = get_dynamic_greeting(AGENT_BRANCH_NAME)
            self._agent_session.say(text=greeting, allow_interruptions=True)
            logger.info("[SUPERVISOR] Delivered greeting via session.say()")


# ============================================================
# ENTRYPOINT
# ============================================================

async def entrypoint(ctx: JobContext):
    logger.info(f"[ENTRYPOINT] Starting supervisor session for room {ctx.room.name}")
    
    # Extract garage_id from room name
    room_name = ctx.room.name
    garage_id = PORTAL_GARAGE_ID
    match = re.match(r'^garage-([a-f0-9-]+)', room_name)
    if match:
        garage_id = match.group(1)
        logger.info(f"[ENTRYPOINT] Extracted garage_id: {garage_id}")
    
    # Refresh configuration for this garage and get agent_mode
    agent_mode = "automate"  # Default
    if garage_id:
        agent_mode = refresh_agent_configuration(garage_id)
        logger.info(f"[ENTRYPOINT] Refreshed configuration for garage: {garage_id}, agent_mode: {agent_mode}")
    
    assist_mode = (agent_mode == "assist")
    if assist_mode:
        logger.info("[ENTRYPOINT] Running in ASSIST mode - limited tool set, no bookings")
    else:
        logger.info("[ENTRYPOINT] Running in AUTOMATE mode - full booking capability")

    # Initialize state and API client
    state = CallState()
    state.call_start_time = time.time()
    gh = GHClient()

    # Create the single supervisor agent
    supervisor = SupervisorAgent(state=state, gh=gh, room_name=room_name, assist_mode=assist_mode)

    # Create session — low-latency config with ElevenLabs TTS
    session = AgentSession(
        vad=silero.VAD.load(),
        turn_detection=MultilingualModel(),
        stt=deepgram.STT(
            model="nova-3",
            language="en-GB",
            interim_results=True,
            smart_format=True,
            punctuate=True,
        ),
        llm="openai/gpt-4.1-mini",
        tts=elevenlabs.TTS(
            model=ELEVEN_TTS_MODEL,
            voice_id=ELEVEN_VOICE_ID,
            voice_settings=elevenlabs.VoiceSettings(
                stability=ELEVEN_STABILITY,
                similarity_boost=ELEVEN_SIMILARITY,
                style=ELEVEN_STYLE,
            ),
        ),
    )

    # Give the agent a reference to the session for session.say()
    supervisor.set_session(session)

    # ── Transcript tracker ─────────────────────────────────────
    # Automatic turn detection via MultilingualModel handles end-of-turn.
    # This handler only tracks recent transcripts for save_caller_name validation.
    _last_final: str = ""

    @session.on("user_input_transcribed")
    def _on_user_transcript(ev):
        nonlocal _last_final
        if not ev.is_final:
            return

        text = ev.transcript.strip()
        if not text:
            return

        # Deduplicate overlapping Deepgram finals.
        prev = _last_final.lower().rstrip(".")
        curr = text.lower().rstrip(".")
        if prev and prev in curr:
            if state.recent_transcripts:
                state.recent_transcripts[-1] = text
            logger.info(f"[TRANSCRIPT] Replaced overlapping final: '{_last_final}' → '{text}'")
        else:
            state.recent_transcripts.append(text)

        _last_final = text

    @session.on("conversation_item_added")
    def _on_conversation_item_added(ev):
        """Capture full agent+customer transcript for GPT summary."""
        try:
            item = getattr(ev, "item", None)
            if item is None:
                return
            role = getattr(item, "role", "") or ""
            text = ""
            if hasattr(item, "text_content"):
                text = item.text_content or ""
            if not text and hasattr(item, "content"):
                content = item.content
                if isinstance(content, str):
                    text = content
                elif isinstance(content, list):
                    parts = []
                    for part in content:
                        if hasattr(part, "text"):
                            parts.append(part.text or "")
                        elif isinstance(part, dict):
                            parts.append(part.get("text", ""))
                    text = " ".join(p for p in parts if p)
            text = text.strip()
            if not text:
                return
            speaker = "agent" if role in ("assistant", "agent") else "customer"
            ts = max(0.0, time.time() - (state.call_start_time or time.time()))
            state.conversation_items.append({"speaker": speaker, "text": text, "timestamp": round(ts, 1)})
            logger.info(f"[TRANSCRIPT] {speaker.capitalize()} speech captured via conversation_item_added: {text[:80]}")
        except Exception as exc:
            logger.warning(f"[TRANSCRIPT] conversation_item_added error: {exc}")

    # Start session
    logger.info("[ENTRYPOINT] Starting session with SupervisorAgent")
    await session.start(
        room=ctx.room,
        agent=supervisor,
    )
    logger.info("[ENTRYPOINT] Session started — supervisor system ready")

    # ── Session isolation: shutdown when caller hangs up ───────
    # Prevents stale conversation context (chat history, turn detector state)
    # from bleeding into the next caller's session in the same room.
    @ctx.room.on("participant_disconnected")
    def _on_caller_left(participant):
        logger.info(
            f"[LIFECYCLE] Caller disconnected: {participant.identity}. "
            "Logging call and shutting down job for clean session isolation."
        )

        async def _shutdown_and_log():
            # Log call to portal before shutdown
            try:
                call_duration = int(time.time() - state.call_start_time) if state.call_start_time else 0

                # ALWAYS use incoming phone number from SIP participant (not what customer says)
                # This ensures accuracy regardless of customer errors when saying their number
                caller_phone = ""
                try:
                    attrs = participant.attributes or {}
                    caller_phone = (
                        attrs.get("sip.phoneNumber") or
                        attrs.get("sip.from") or
                        ""
                    )
                    # Fall back: parse identity like sip_+447841422472
                    if not caller_phone and participant.identity.startswith("sip_"):
                        caller_phone = participant.identity[4:]  # strip 'sip_'
                    if caller_phone:
                        logger.info(f"[PORTAL] Using incoming caller phone from SIP: {caller_phone}")
                    else:
                        # Only if we can't get SIP phone, use what customer said
                        caller_phone = state.contact_phone or ""
                        logger.info(f"[PORTAL] No SIP phone available, using customer-provided: {caller_phone}")
                except Exception:
                    caller_phone = state.contact_phone or ""
                    logger.warning(f"[PORTAL] Error extracting SIP phone, using customer-provided: {caller_phone}")

                # Build transcript: prefer conversation_items (full agent+customer turns captured in real-time)
                # Fall back to synthetic transcript from recent_transcripts if no conversation_items
                base_ts = state.call_start_time or time.time()
                if state.conversation_items:
                    transcript = state.conversation_items
                    # Ensure at least one agent entry exists
                    if not any(e.get("speaker") == "agent" for e in transcript):
                        transcript = [{"speaker": "agent", "text": "Hello, how can I help you today?", "timestamp": 0}] + transcript
                    logger.info(f"[PORTAL] Using {len(transcript)} conversation_items for transcript")
                else:
                    # Fallback: synthetic transcript from customer utterances
                    logger.info("[PORTAL] No conversation_items — building synthetic transcript from recent_transcripts")
                    transcript = []
                    transcript.append({"speaker": "agent", "text": "Hello, how can I help you today?", "timestamp": 0})
                    for i, text in enumerate(state.recent_transcripts or [], start=1):
                        transcript.append({"speaker": "customer", "text": text, "timestamp": i * 5})
                    # Append key structured events as agent turns
                    offset = len(transcript) * 5
                    if state.vrn:
                        transcript.append({"speaker": "agent", "text": f"I have your vehicle registration as {state.vrn}.", "timestamp": offset})
                        offset += 5
                    if state.booking_date:
                        transcript.append({"speaker": "agent", "text": f"Booking confirmed for {state.booking_date} at {state.booking_time}.", "timestamp": offset})

                # Generate GPT summary (falls back to state-based if LLM unavailable)
                summary = await generate_call_summary(transcript, state)

                # Determine call type
                call_type = "unknown"
                if state.intent == "booking" and state.booking_date:
                    call_type = "booking"
                elif state.intent == "quote":
                    call_type = "quote"
                elif state.intent == "message":
                    call_type = "message"
                elif state.intent == "vehicle_update":
                    call_type = "vehicle_update"

                # Build booking details if applicable
                booking_details = ""
                if state.booking_date:
                    booking_parts = []
                    booking_parts.append(f"Date: {state.booking_date}")
                    if state.booking_time:
                        booking_parts.append(f"Time: {state.booking_time}")
                    if state.service_selected_name:
                        booking_parts.append(f"Service: {state.service_selected_name}")
                    if state.service_price:
                        booking_parts.append(f"Price: {state.service_price}")
                    booking_details = ", ".join(booking_parts)

                # metrics must be a non-empty object
                metrics = {
                    "duration_seconds": call_duration,
                    "intent": state.intent or "unknown",
                    "vrn_captured": bool(state.vrn),
                    "booking_confirmed": state.step == Step.CONFIRMED,
                }

                logger.info(f"[PORTAL] Extracted caller phone: {caller_phone}")
                logger.info(f"[PORTAL] Transcript entries: {len(transcript)}, summary: {summary[:80]}")

                # Log to portal
                await log_call_to_portal(
                    garage_id=garage_id,
                    room_name=room_name,
                    duration_seconds=call_duration,
                    transcript=transcript,
                    summary=summary,
                    customer_name=f"{state.customer_name_first} {state.customer_name_last}".strip() or "Unknown",
                    customer_phone=caller_phone,
                    registration_number=state.vrn,
                    confirmed_booking=state.step == Step.CONFIRMED,
                    booking_details=booking_details,
                    call_type=call_type,
                    metrics=metrics,
                )
            except Exception as e:
                logger.error(f"[PORTAL] Failed to log call: {e}")

            ctx.shutdown("caller_disconnected")

        asyncio.create_task(_shutdown_and_log())


# ============================================================
# CLI
# ============================================================

if __name__ == "__main__":
    cli.run_app(
        WorkerOptions(
            entrypoint_fnc=entrypoint,
            worker_type=WorkerType.ROOM,
            agent_name="receptionmate-agent-v3",
        )
    )
