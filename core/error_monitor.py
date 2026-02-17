from __future__ import annotations

import asyncio
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import aiohttp


@dataclass(frozen=True)
class ErrorMonitorConfig:
    discord_webhook_url: str = ""
    excel_path: Path | None = None
    timezone_display: str = "Europe/London"


class ErrorMonitor:
    """Non-fatal error reporting via Discord and Excel log."""

    def __init__(self, config: ErrorMonitorConfig, logger: logging.Logger | None = None) -> None:
        self._config = config
        self._logger = logger or logging.getLogger(__name__)

    async def report_error(
        self,
        *,
        error_msg: str,
        agent_name: str,
        room_name: str = "",
        error_type: str = "",
        extra: Optional[dict] = None,
    ) -> None:
        tasks = [self._log_to_excel(error_msg, agent_name, room_name, error_type)]
        if self._config.discord_webhook_url:
            tasks.append(
                self._send_discord_alert(
                    error_msg=error_msg,
                    agent_name=agent_name,
                    room_name=room_name,
                    extra=extra,
                )
            )
        await asyncio.gather(*tasks, return_exceptions=True)

    async def _send_discord_alert(
        self,
        *,
        error_msg: str,
        agent_name: str,
        room_name: str,
        extra: Optional[dict],
    ) -> None:
        from zoneinfo import ZoneInfo

        embed = {
            "title": f"ReceptionMate Error - {agent_name}",
            "description": error_msg[:2000],
            "color": 0xFF0000,
            "fields": [
                {"name": "Agent", "value": agent_name or "N/A", "inline": True},
                {"name": "Room", "value": room_name or "N/A", "inline": True},
                {
                    "name": "Time",
                    "value": datetime.now(ZoneInfo(self._config.timezone_display)).strftime("%Y-%m-%d %H:%M:%S"),
                    "inline": True,
                },
            ],
            "footer": {"text": "ReceptionMate Supervisor System"},
        }
        if extra:
            for key, value in extra.items():
                embed["fields"].append({"name": key, "value": str(value)[:1024], "inline": False})

        try:
            timeout = aiohttp.ClientTimeout(total=5)
            async with aiohttp.ClientSession(timeout=timeout) as session:
                async with session.post(
                    self._config.discord_webhook_url,
                    json={"embeds": [embed]},
                ) as resp:
                    if resp.status not in (200, 204):
                        self._logger.warning("[ErrorMonitor] Discord webhook returned %s", resp.status)
        except Exception as exc:  # pragma: no cover
            self._logger.warning("[ErrorMonitor] Failed to send Discord alert: %s", exc)

    async def _log_to_excel(self, error_msg: str, agent_name: str, room_name: str, error_type: str) -> None:
        if not self._config.excel_path:
            return

        async def _write() -> None:
            try:
                from openpyxl import load_workbook, Workbook  # type: ignore
            except ImportError:  # pragma: no cover
                self._logger.warning("[ErrorMonitor] openpyxl not installed; skipping Excel logging")
                return

            path = self._config.excel_path
            if not path:
                return

            if path.exists():
                wb = load_workbook(str(path))
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Error Log"
                ws.append(["Timestamp", "Agent", "Room", "Error Type", "Error Message"])

            from zoneinfo import ZoneInfo

            ws.append([
                datetime.now(ZoneInfo(self._config.timezone_display)).strftime("%Y-%m-%d %H:%M:%S"),
                agent_name,
                room_name or "N/A",
                error_type or "general",
                error_msg[:5000],
            ])
            wb.save(str(path))

        await asyncio.to_thread(_write)
